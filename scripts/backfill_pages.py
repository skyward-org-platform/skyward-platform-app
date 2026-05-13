"""Generic page backfill: WQA workbook URL Triage tab -> page table.

Reads column positions by HEADER NAME so it works across WQA workbook
versions (phil-lasry's 41-col layout, minibushire's 30-col layout, etc.).

Usage:
    python backfill_pages.py <property_slug> <workbook_path> [--audit-date YYYY-MM-DD]

Examples:
    python backfill_pages.py buscharter delivery/tna/buscharter/phase-1-wqa/BusCharter-Website-Quality-Audit-v2-2026-04-27.xlsx
    python backfill_pages.py tnabushire delivery/tna/tnabushire/phase-1-wqa/tnabushire-Website-Quality-Audit-v2-2026-04-27.xlsx --audit-date 2026-04-27
"""
from __future__ import annotations

import argparse
import os
import re
import sys
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from supabase import create_client

load_dotenv("/Users/paulskirbe/agency/.env")

TRIAGE_SHEET_NAME = "URL Triage"

# Header name -> column key in page table. Title/H1/RedirectFinalURL are optional.
REQUIRED_HEADERS = {"URL": "url", "Action": "audit_action"}
OPTIONAL_HEADERS = {
    "Logic": "audit_notes",
    "Status Code": "status_code",
    "Category": "page_type",
    "Title": "title",
    "H1": "h1",
    "Redirect Final URL": "audit_target_url",
}

ACTION_VERBS = {
    "optimize": "optimize",
    "restore": "restore",
    "redirect": "redirect",
    "consolidate": "consolidate",
    "remove": "remove",
    "delete": "remove",
    "keep": "keep",
    "no action": "no_action",
    "no_action": "no_action",
    "review": "undecided",
}


def _normalize_action(raw: object) -> str:
    """Map workbook Action cell to schema enum.

    Handles both plain values ("Optimize") and v2-style parenthetical forms
    ("Optimize (revenue-critical)", "No Action (system URL)") by stripping
    the parenthetical and matching the leading verb.
    """
    if raw is None:
        return "undecided"
    s = str(raw).strip()
    if not s:
        return "undecided"
    # Strip trailing parenthetical reason. "No Action (system URL)" -> "No Action".
    if "(" in s:
        s = s.split("(", 1)[0].strip()
    return ACTION_VERBS.get(s.lower(), "undecided")


def _clean(val: object) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s and s.lower() not in ("none found", "none") else None


def _admin_client():
    return create_client(os.environ["SUPABASE_URL"], os.environ["SUPABASE_SERVICE_ROLE_KEY"])


def _resolve_columns(header: tuple) -> dict[str, int]:
    """Map our column keys to their positions in this workbook's header row."""
    positions: dict[str, int] = {}
    for i, h in enumerate(header):
        if h in REQUIRED_HEADERS:
            positions[REQUIRED_HEADERS[h]] = i
        elif h in OPTIONAL_HEADERS:
            positions[OPTIONAL_HEADERS[h]] = i
    missing = set(REQUIRED_HEADERS.values()) - set(positions)
    if missing:
        raise SystemExit(f"Workbook is missing required columns: {missing}")
    return positions


def iter_triage_rows(workbook_path: Path):
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    if TRIAGE_SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"No '{TRIAGE_SHEET_NAME}' tab in {workbook_path.name}")
    ws = wb[TRIAGE_SHEET_NAME]
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    cols = _resolve_columns(header)

    for row in ws.iter_rows(min_row=2, values_only=True):
        url_val = row[cols["url"]] if cols["url"] < len(row) else None
        if not url_val:
            continue
        url = str(url_val).strip()
        if not url.startswith("http"):
            continue

        raw_action = row[cols["audit_action"]] if cols["audit_action"] < len(row) else None
        audit_action = _normalize_action(raw_action)

        record: dict[str, object | None] = {
            "url": url,
            "audit_action": audit_action,
        }
        for k in ("audit_notes", "page_type", "title", "h1", "audit_target_url"):
            if k in cols and cols[k] < len(row):
                record[k] = _clean(row[cols[k]])
        if "status_code" in cols and cols["status_code"] < len(row):
            sc = row[cols["status_code"]]
            record["status_code"] = int(sc) if sc not in (None, "") else None

        yield record


def _infer_audit_date(workbook_path: Path) -> str:
    """Pull YYYY-MM-DD from the filename if present, else today."""
    m = re.search(r"(\d{4}-\d{2}-\d{2})", workbook_path.name)
    if m:
        return f"{m.group(1)}T00:00:00Z"
    from datetime import datetime, timezone
    return datetime.now(timezone.utc).date().isoformat() + "T00:00:00Z"


def run(property_slug: str, workbook_path: Path, audit_date: str | None) -> None:
    db = _admin_client()
    prop = db.table("property").select("id, name").eq("slug", property_slug).single().execute().data
    if not prop:
        raise SystemExit(f"No property found with slug={property_slug!r}")
    property_id = prop["id"]
    audit_decided_at = audit_date or _infer_audit_date(workbook_path)

    rows = list(iter_triage_rows(workbook_path))
    print(f"[{property_slug}] read {len(rows)} pages from {workbook_path.name}")
    print(f"[{property_slug}] audit_decided_at={audit_decided_at}")

    batch_size = 100
    inserted = 0
    for i in range(0, len(rows), batch_size):
        batch = [
            {
                **r,
                "property_id": property_id,
                "audit_decided_by": "import:wqa_workbook",
                "audit_decided_at": audit_decided_at,
            }
            for r in rows[i : i + batch_size]
        ]
        db.table("page").upsert(batch, on_conflict="property_id,url").execute()
        inserted += len(batch)
        print(f"[{property_slug}]   upserted {inserted}/{len(rows)}")

    # Quick distribution summary
    from collections import Counter
    dist = Counter(r["audit_action"] for r in rows)
    print(f"[{property_slug}] action distribution: {dict(dist)}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("property_slug")
    ap.add_argument("workbook_path", type=Path)
    ap.add_argument("--audit-date", help="YYYY-MM-DD audit decision date; defaults to date in filename or today")
    args = ap.parse_args()
    audit_iso = f"{args.audit_date}T00:00:00Z" if args.audit_date else None
    run(args.property_slug, args.workbook_path, audit_iso)


if __name__ == "__main__":
    main()
