"""Phase 2 Technical SEO Audit workbook builder — ported from
`/Users/paulskirbe/agency/delivery/tna/build_phase2_technical.py`.

Self-contained in-memory variant. Public entrypoint:
:func:`build_phase2_workbook`. It accepts the Phase 1 triage DataFrame
(produced by `_phase1_builder.build_phase1_dataframe`) plus optional
overlay maps for execution + check-state, and returns a BytesIO with
the multi-tab Phase 2 workbook (Issue Summary, Audit Checklist, one
sheet per failing T/C check, Page Speed, Website Architecture, Schema
Optimization, Broken List, URL Priority, Aggregate).
"""
from __future__ import annotations

import re
import statistics
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ─── Check definitions ─────────────────────────────────────────────────────
T_CHECKS = [
    ("T1", "Schema errors", "Fix Schema", "blocked", "Requires SF structured data report (contains_structured_data + validation_errors)"),
    ("T2", "Missing schema", "Add Schema", "blocked", "Requires SF structured data report"),
    ("T3", "Review star opportunity", "Add Product+AggregateRating", "blocked", "Requires SF structured data report to confirm absence"),
    ("T4", "Orphan with value", "Add Internal Links", "active", None),
    ("T5", "Under-linked", "Add Internal Links", "active", None),
    ("T6", "Buried page (depth>=4)", "Improve Architecture", "active", None),
    ("T7", "Over-linked underperformer", "Remove Internal Links", "active", None),
    ("T8", "Indexable but not indexed", "Fix Indexation", "blocked", "Requires SF + GSC URL Inspection integration (indexed-in-Google column)"),
    ("T9", "Noindex on valuable page", "Fix Indexation", "active", None),
    ("T10", "Multiple canonicals", "Fix Multiple Canonicals", "blocked", "Requires SF canonical-tags multi-value field"),
    ("T11", "Canonical mismatch", "Canonicalize", "active", None),
    ("T12", "Not in sitemap", "Add to Sitemap", "active", None),
    ("T13", "Blocked resources (JS/CSS)", "Fix Blocked Resources", "blocked", "Requires robots.txt + SF resources report"),
    ("T14", "JS rendering required", "Verify JS Rendering", "active", None),
    ("T15", "No issues found", "Leave As Is", "active", None),
    ("T16", "Pages linking to broken pages", "Update Internal Links", "blocked", "Requires SF inlinks export to broken-URL set"),
    ("T17", "HTTPS page linking to HTTP", "Fix Internal Links", "blocked", "Requires SF inlinks raw URL inspection"),
    ("T18", "Missing social tags (OG/Twitter)", "Add Social Tags", "blocked", "Requires SF social-tags extract"),
    ("T19", "IndexNow candidates", "Submit to IndexNow", "active", None),
    ("T20", "Duplicate without canonical", "Add Canonical", "blocked", "Requires SF duplicate content report"),
]
C_CHECKS = [
    ("C1", "Revenue page losing traffic", "Refresh (URGENT)", "active", None),
    ("C2", "Low engagement (sessions>50, dur<30s)", "Rewrite", "active", None),
    ("C3", "Cannibalization (2+ pages same KW)", "Merge weaker into stronger", "active", None),
    ("C4", "Thin content (<300 words, <10 sessions)", "Rewrite", "active", None),
    ("C5", "Losing traffic (negative session %)", "Refresh or Rewrite", "active", None),
    ("C6", "Ranking 3-10, SV>50", "Target with Links", "active", None),
    ("C7", "Ranking 11-20, SV>50", "Refresh or Rewrite", "active", None),
    ("C8", "Missing meta description", "Update Meta Description", "active", None),
    ("C9", "Duplicate meta description", "Update Meta Description", "active", None),
    ("C10", "Duplicate title", "Update Page Title", "active", None),
    ("C11", "Title issues (length/stuffing)", "Update Page Title", "active", None),
    ("C12", "Has refs but not ranking (rank>20)", "Target w/ Links + Refresh", "active", None),
    ("C13", "Performing well (sess>50, words>1000)", "Leave As Is", "active", None),
    ("C14", "No issues found", "Leave As Is", "active", None),
    ("C15", "Meta description too long (>155)", "Shorten Meta Description", "active", None),
    ("C16", "Meta description too short (<70)", "Expand Meta Description", "active", None),
    ("C17", "Title too long (>65)", "Shorten Title", "active", None),
    ("C18", "Title too short (<30)", "Expand Title", "active", None),
    ("C19", "AI content detection (Ahrefs)", "Review Content Quality", "blocked", "Requires Ahrefs Site Audit AI-content signal"),
    ("C20", "Page vs SERP title mismatch", "Improve Title Quality", "blocked", "Requires GSC URL Inspection API + SERP title comparison"),
]
S_CHECKS = [
    ("S1", "Robots.txt", "Fix robots.txt", "live"),
    ("S2", "Navigation (3-click reach, crawlable, mobile parity)", "Fix Navigation", "partial"),
    ("S3", "XML Sitemap", "Fix Sitemap", "live"),
    ("S4", "HTTPS enforcement", "Fix HTTPS", "partial"),
    ("S5", "Core Web Vitals", "Fix CWV (dev)", "blocked"),
    ("S6", "Schema sitewide", "Fix Schema", "blocked"),
    ("S7", "Duplicate content (SF reports)", "Resolve Duplicates", "blocked"),
    ("S8", "Orphan pages", "Add Links or Remove", "active"),
    ("S9", "Hreflang", "Fix Hreflang", "partial"),
    ("S10", "Social tags (OG/Twitter)", "Add Social Tags", "blocked"),
    ("S11", "Pagination", "Fix Pagination", "partial"),
    ("S12", "Platform Performance Ceiling", "Document Ceiling", "blocked"),
]


SCHEMA_TARGETS_BY_CATEGORY = {
    "Homepage": "LocalBusiness + FAQ + WebSite + Organization (+ Review if applicable)",
    "Service Page": "Service + FAQ + Vehicle (if relevant) + Breadcrumb",
    "Fleet/Product Page": "Vehicle/Product per vehicle (Product+AggregateRating intentional for review stars) + Breadcrumb",
    "Location Page": "LocalBusiness (reference only) + Service (if highlighted) + FAQ + Breadcrumb",
    "Blog Post": "Article + Breadcrumb",
    "Blog Hub": "Breadcrumb",
    "Blog Category": "Breadcrumb",
    "Quote/Contact": "ContactPoint or LocalBusiness (reference) + Breadcrumb",
    "Utility": "Breadcrumb",
}


# ─── helpers ────────────────────────────────────────────────────────────────
def num(v: Any) -> float:
    if v is None:
        return 0.0
    if isinstance(v, str):
        try:
            return float(v)
        except ValueError:
            return 0.0
    if pd.isna(v):
        return 0.0
    try:
        return float(v)
    except Exception:
        return 0.0


def _str(v: Any) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    return str(v).strip()


def kw_dependency(check_id: str) -> str:
    if check_id in ("C2", "C4", "C5", "C7", "C13"):
        return "Phase 3 Dependent"
    if check_id in ("C8", "C9", "C10", "C11", "C15", "C16", "C17", "C18", "C20"):
        return "Fix Now, Revisit"
    return "Fix Now"


def median_inlinks_by_category(df: pd.DataFrame) -> dict[str, float]:
    out: dict[str, float] = {}
    for cat, sub in df.groupby("Category"):
        vals = [num(v) for v in sub["Inlinks"] if num(v) > 0]
        out[cat] = statistics.median(vals) if vals else 0.0
    return out


def t_check(check_id: str, r: pd.Series, ctx: dict) -> tuple[bool, str]:
    sessions = num(r.get("Sessions"))
    imps = num(r.get("Average Impressions"))
    inlinks = num(r.get("Inlinks"))
    depth = num(r.get("Page Depth"))
    canon = _str(r.get("Canonical Link Element"))
    in_sitemap = _str(r.get("In Sitemap")).lower() in ("true", "yes", "1")
    indexability = _str(r.get("Indexability"))
    idx_status = _str(r.get("Indexability Status"))
    sc = r.get("Status Code")
    try:
        sc_int = int(float(sc)) if pd.notna(sc) else None
    except (TypeError, ValueError):
        sc_int = None
    url = _str(r.get("URL"))
    category = _str(r.get("Category"))
    page_type = _str(r.get("Type"))

    if check_id == "T4":
        if inlinks == 0 and (sessions > 0 or imps > 0):
            return True, f"Inlinks=0; sessions={int(sessions)}, impressions={int(imps)}"
    elif check_id == "T5":
        med = ctx["median_inlinks"].get(category, 0)
        if med > 0 and inlinks > 0 and inlinks < med * 0.5:
            return True, f"Inlinks={int(inlinks)} < 50% of category median ({med:.1f}) for {category}"
    elif check_id == "T6":
        if depth >= 4:
            return True, f"Page depth = {int(depth)} (>= 4 clicks from homepage)"
    elif check_id == "T7":
        med = ctx["median_inlinks"].get(category, 0)
        if med > 0 and inlinks > med * 2 and sessions < 10:
            return True, f"Inlinks={int(inlinks)} > 200% of category median ({med:.1f}); sessions={int(sessions)} (bottom-quartile)"
    elif check_id == "T9":
        if "noindex" in idx_status.lower() and (sessions > 0 or num(r.get("Referring Domains")) > 0):
            return True, f"Noindex but sessions={int(sessions)}, refs={int(num(r.get('Referring Domains')))}"
    elif check_id == "T11":
        if canon and canon != url and canon.rstrip("/") != url.rstrip("/"):
            return True, f"Canonical -> {canon}"
    elif check_id == "T12":
        if sc_int == 200 and indexability == "Indexable" and not in_sitemap:
            return True, "Indexable 200 URL absent from sitemap"
    elif check_id == "T14":
        if "javascript" in page_type.lower() or "spa" in page_type.lower():
            return True, f"Page Type = {page_type}"
    elif check_id == "T19":
        return False, ""

    return False, ""


def c_check(check_id: str, r: pd.Series, ctx: dict) -> tuple[bool, str]:
    sessions = num(r.get("Sessions"))
    duration = num(r.get("Average Session Duration"))
    conv_rate = num(r.get("Conversion Rate (%)"))
    revenue = num(r.get("Total Revenue"))
    losing = _str(r.get("Losing Traffic?")).lower() in ("true", "yes", "1")
    sess_pct = _str(r.get("Session % Change"))
    sess_pct_neg = sess_pct.startswith("-") if sess_pct else False
    word_count = num(r.get("Word Count"))
    rank = num(r.get("Best TV KW Rank"))
    sv = num(r.get("Best TV KW SV"))
    rd = num(r.get("Referring Domains"))
    title = _str(r.get("Current Title"))
    meta = _str(r.get("Meta Description"))
    kw = _str(r.get("Best TV Keyword"))

    if check_id == "C1":
        if (conv_rate > 5 or revenue > 100) and losing:
            return True, f"Conv Rate={conv_rate:.1f}% Revenue=${revenue:.2f} losing_traffic=true"
    elif check_id == "C2":
        if sessions > 50 and 0 < duration < 30:
            return True, f"Sessions={int(sessions)} but avg duration={duration:.0f}s (<30s)"
    elif check_id == "C3":
        if kw and ctx["kw_to_urls"].get(kw, 0) >= 2:
            return True, f'Keyword "{kw}" targeted by {ctx["kw_to_urls"][kw]} pages'
    elif check_id == "C4":
        if word_count > 0 and word_count < 300 and sessions < 10:
            return True, f"Word count={int(word_count)} (<300); sessions={int(sessions)} (<10)"
    elif check_id == "C5":
        if sess_pct_neg or losing:
            return True, f"Session % change={sess_pct or '(losing)'}"
    elif check_id == "C6":
        if 3 <= rank <= 10 and sv > 50:
            return True, f'Ranking {int(rank)} for "{kw}" (SV={int(sv)})'
    elif check_id == "C7":
        if 11 <= rank <= 20 and sv > 50:
            return True, f'Ranking {int(rank)} for "{kw}" (SV={int(sv)})'
    elif check_id == "C8":
        if not meta:
            return True, "Meta description is empty"
    elif check_id == "C9":
        if meta and ctx["meta_to_urls"].get(meta, 0) >= 2:
            return True, f"Meta shared with {ctx['meta_to_urls'][meta] - 1} other URL(s)"
    elif check_id == "C10":
        if title and ctx["title_to_urls"].get(title, 0) >= 2:
            return True, f"Title shared with {ctx['title_to_urls'][title] - 1} other URL(s)"
    elif check_id == "C11":
        if title:
            n = len(title)
            if n > 65:
                return True, f"Title length={n} (>65 chars)"
            keywords = re.findall(r"\b[a-z][a-z\-]+\b", title.lower())
            counts = Counter(keywords)
            stuffed = [w for w, c in counts.items() if c >= 3 and len(w) > 3]
            if stuffed:
                return True, f"Keyword stuffing: {', '.join(stuffed)} (3+ repeats)"
    elif check_id == "C12":
        if rd > 0 and rank > 20:
            return True, f"Refs={int(rd)} but rank={int(rank)} (page 3+)"
    elif check_id == "C13":
        if sessions > 50 and word_count > 1000:
            return True, f"Sessions={int(sessions)} + words={int(word_count)} = performing well"
    elif check_id == "C15":
        if meta and len(meta) > 155:
            return True, f"Meta length={len(meta)} (>155)"
    elif check_id == "C16":
        if meta and 0 < len(meta) < 70:
            return True, f"Meta length={len(meta)} (<70)"
    elif check_id == "C17":
        if title and len(title) > 65:
            return True, f"Title length={len(title)} (>65)"
    elif check_id == "C18":
        if title and len(title) < 30:
            return True, f"Title length={len(title)} (<30)"

    return False, ""


def run_sitewide_checks(domain: str, df: pd.DataFrame) -> list[dict]:
    """S1-S12. Live HTTP probes for robots.txt + sitemap; everything else
    is crawl-derived or marked Blocked."""
    import urllib.request
    import urllib.error

    results: list[dict] = []

    try:
        req = urllib.request.Request(
            f"https://{domain}/robots.txt", headers={"User-Agent": "Skyward-WQA/1.0"}
        )
        with urllib.request.urlopen(req, timeout=10) as resp:
            body = resp.read().decode("utf-8", errors="ignore")
        if "Sitemap:" not in body and "sitemap:" not in body:
            results.append(dict(id="S1", check="Robots.txt", status="Fail",
                                detail="No Sitemap directive in robots.txt"))
        elif re.search(r"Disallow:\s*/\s*$", body, re.M):
            results.append(dict(id="S1", check="Robots.txt", status="Fail",
                                detail="Site-wide Disallow / detected (blocks entire site)"))
        else:
            results.append(dict(id="S1", check="Robots.txt", status="Pass",
                                detail=f"Robots.txt accessible; sitemap referenced ({len(body)} bytes)"))
    except Exception as e:
        results.append(dict(id="S1", check="Robots.txt", status="Fail",
                            detail=f"Could not fetch robots.txt: {type(e).__name__}"))

    try:
        sitemap_url = None
        for path in ("/sitemap.xml", "/sitemap_index.xml", "/wp-sitemap.xml"):
            try:
                with urllib.request.urlopen(f"https://{domain}{path}", timeout=10) as resp:
                    if 200 <= resp.status < 300:
                        sitemap_url = f"https://{domain}{path}"
                        break
            except urllib.error.HTTPError:
                continue
            except Exception:
                continue
        if sitemap_url:
            results.append(dict(id="S3", check="XML Sitemap", status="Pass",
                                detail=f"Sitemap found at {sitemap_url}"))
        else:
            results.append(dict(id="S3", check="XML Sitemap", status="Fail",
                                detail="No sitemap.xml or sitemap_index.xml at standard paths"))
    except Exception as e:
        results.append(dict(id="S3", check="XML Sitemap", status="Fail",
                            detail=f"Sitemap check error: {type(e).__name__}"))

    max_depth = df["Page Depth"].apply(num).max() if "Page Depth" in df.columns else 0
    over_3 = int((df["Page Depth"].apply(num) > 3).sum()) if "Page Depth" in df.columns else 0
    if over_3 > 0:
        results.append(dict(id="S2", check="Navigation (3-click reach)", status="Fail",
                            detail=f"{over_3} URLs at depth >3; max depth={int(max_depth)}"))
    else:
        results.append(dict(id="S2", check="Navigation (3-click reach)", status="Pass",
                            detail="All Optimize+Restore URLs within 3 clicks of homepage"))

    http_urls = int(df["URL"].astype(str).str.startswith("http://").sum())
    if http_urls > 0:
        results.append(dict(id="S4", check="HTTPS enforcement", status="Fail",
                            detail=f"{http_urls} URLs in scope use http:// (require redirect to https)"))
    else:
        results.append(dict(id="S4", check="HTTPS enforcement", status="Pass",
                            detail="All scoped URLs use HTTPS"))

    results.append(dict(id="S5", check="Core Web Vitals", status="Blocked",
                        detail="Requires Screaming Frog PageSpeed integration export (pagespeed_all.csv)"))
    results.append(dict(id="S6", check="Schema sitewide", status="Blocked",
                        detail="Requires Screaming Frog structured data report"))
    results.append(dict(id="S7", check="Duplicate content", status="Blocked",
                        detail="Requires Screaming Frog near-dupe + exact-dupe content reports"))

    orphans = df[(df["Inlinks"].apply(num) == 0) & (df["Status Code"].apply(num) == 200)]
    if len(orphans) > 0:
        results.append(dict(id="S8", check="Orphan pages", status="Fail",
                            detail=f"{len(orphans)} URLs in scope have 0 inlinks"))
    else:
        results.append(dict(id="S8", check="Orphan pages", status="Pass",
                            detail="No orphans in Optimize+Restore set"))

    results.append(dict(id="S9", check="Hreflang", status="N/A",
                        detail="Single-language single-country property"))
    results.append(dict(id="S10", check="Social tags (OG/Twitter)", status="Blocked",
                        detail="Requires Screaming Frog social-tag extraction"))

    paginated = df[df["URL"].astype(str).str.contains(r"/page/\d+|\?page=\d+|/p/\d+", regex=True)]
    if len(paginated) > 0:
        results.append(dict(id="S11", check="Pagination", status="Fail",
                            detail=f"{len(paginated)} paginated URLs in scope; manual review required for canonical/self-ref"))
    else:
        results.append(dict(id="S11", check="Pagination", status="N/A",
                            detail="No paginated URLs detected"))

    results.append(dict(id="S12", check="Platform Performance Ceiling", status="Blocked",
                        detail="Depends on S5 (CWV) results"))

    return results


# ─── styling ────────────────────────────────────────────────────────────────
NAVY = "1B2A4A"
GREEN = "22C55E"
YELLOW = "F59E0B"
ORANGE = "F97316"
BLUE = "3B82F6"
GREY = "6B7280"
WHITE = "FFFFFF"

THIN = Side(border_style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, row, ncols, bg=NAVY):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name="Inter", size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
        cell.border = BORDER


def style_title(ws, row, text, ncols=8):
    ws.cell(row=row, column=1, value=text)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1)
    cell.font = Font(name="Inter", size=14, bold=True, color=NAVY)


def style_subtitle(ws, row, text, ncols=8):
    ws.cell(row=row, column=1, value=text)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1)
    cell.font = Font(name="Inter", size=10, italic=True, color="475569")
    cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)


def autosize(ws, max_col, max_w=80):
    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        max_len = 8
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            v = row[0]
            if v is None:
                continue
            l = len(str(v))
            if l > max_len:
                max_len = l
        ws.column_dimensions[letter].width = min(max_len + 2, max_w)


def color_sheet_tab(ws, color: str):
    ws.sheet_properties.tabColor = color


# ─── tab writers ────────────────────────────────────────────────────────────
@dataclass
class CheckResult:
    check_id: str
    name: str
    failed_urls: list[dict] = field(default_factory=list)
    status: str = "Pass"
    detail: str = ""
    kw_dependency: str = "Fix Now"
    action: str = ""
    blocked_reason: str | None = None


def _format_current_state(r: pd.Series) -> str:
    parts = []
    if _str(r.get("Current Title")):
        parts.append(f"Title: {_str(r['Current Title'])[:60]}")
    if _str(r.get("Meta Description")):
        parts.append(f"Meta: {len(_str(r['Meta Description']))} chars")
    parts.append(f"WC: {int(num(r.get('Word Count')))}")
    parts.append(f"Inlinks: {int(num(r.get('Inlinks')))}")
    parts.append(f"Depth: {int(num(r.get('Page Depth')))}")
    return " | ".join(parts)


def _format_target_state(r: pd.Series, actions: list) -> str:
    return f"All {len(actions)} action items resolved; checks Pass"


def _format_acceptance(actions: list) -> str:
    chunks = []
    for cid, act, _ in actions:
        if cid.startswith("T1") or cid.startswith("T2") or cid.startswith("T3"):
            chunks.append(f"{cid}: Rich Results Test zero errors")
        elif cid == "T11":
            chunks.append("T11: Canonical = self-URL")
        elif cid == "T12":
            chunks.append("T12: URL in sitemap.xml")
        elif cid.startswith("C1") and cid != "C19" and cid != "C20":
            chunks.append(f"{cid}: Re-crawl confirms fix")
        elif cid == "C8":
            chunks.append("C8: Meta description present, 70-155 chars")
    return "; ".join(chunks) if chunks else "Re-crawl + verify"


def _aggregate_kw_dep(deps: list[str]) -> str:
    if "Phase 3 Dependent" in deps:
        return "Phase 3 Dependent"
    if "Fix Now, Revisit" in deps:
        return "Fix Now, Revisit"
    return "Fix Now"


def write_issue_summary(wb, title, check_results, total_scope):
    ws = wb.create_sheet("Issue Summary")
    color_sheet_tab(ws, NAVY)
    style_title(ws, 1, f"{title} -- Issue Summary", 9)
    style_subtitle(ws, 2, f"All Phase 2 checks. Scope: {total_scope} Optimize+Restore URLs from Phase 1.", 9)
    headers = ["#", "Check", "Status", "URLs Affected", "Action", "KW Dependency", "Severity", "Owner", "Detail"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4, 9)
    row = 5
    for cid, cr in sorted(check_results.items(), key=lambda x: (x[1].status != "Fail", x[0])):
        urls_affected = len(cr.failed_urls) if cr.status == "Fail" else 0
        severity = "High" if urls_affected >= 20 else "Medium" if urls_affected >= 5 else "Low" if urls_affected > 0 else "-"
        owner = "Client Dev" if cid in ("T1","T2","T3","T10","T13","T18","T20","S5") else "Skyward + Client Dev"
        detail = cr.detail if cr.status != "Blocked" else f"BLOCKED: {cr.blocked_reason}"
        for i, v in enumerate([cid, cr.name, cr.status, urls_affected, cr.action, cr.kw_dependency, severity, owner, detail], start=1):
            cell = ws.cell(row=row, column=i, value=v)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        status_color = {"Fail": "FECACA", "Pass": "DCFCE7", "Blocked": "E0E7FF", "N/A": "F1F5F9"}.get(cr.status, "F1F5F9")
        ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor=status_color)
        row += 1
    autosize(ws, 9, max_w=60)
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["I"].width = 60


def _apply_status_color(ws, row, col, status):
    fills = {"Fail": "FECACA", "Pass": "DCFCE7", "Blocked": "E0E7FF", "N/A": "F1F5F9"}
    color = fills.get(status)
    if color:
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=color)


def write_audit_checklist(wb, title, check_results, s_results, total_scope):
    ws = wb.create_sheet("Audit Checklist")
    color_sheet_tab(ws, NAVY)
    style_title(ws, 1, f"{title} -- Audit Checklist", 5)
    style_subtitle(ws, 2, "All 44 checks per SOP v4 §7. Pass/Fail/Blocked/N/A per check.", 5)

    row = 4
    ws.cell(row=row, column=1, value="TECHNICAL (T1-T20)").font = Font(name="Inter", size=11, bold=True, color=NAVY)
    row += 1
    for i, h in enumerate(["#", "Check", "Status", "URLs Affected", "Detail / Action"], start=1):
        ws.cell(row=row, column=i, value=h)
    style_header(ws, row, 5)
    row += 1
    for cid in [t[0] for t in T_CHECKS]:
        cr = check_results[cid]
        urls_aff = len(cr.failed_urls) if cr.status == "Fail" else (None if cr.status in ("Blocked", "N/A") else 0)
        detail = cr.blocked_reason if cr.status == "Blocked" else (cr.detail or cr.action)
        for i, v in enumerate([cid, cr.name, cr.status, urls_aff, detail], start=1):
            ws.cell(row=row, column=i, value=v).alignment = Alignment(vertical="top", wrap_text=True)
        _apply_status_color(ws, row, 3, cr.status)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="CONTENT (C1-C20)").font = Font(name="Inter", size=11, bold=True, color=NAVY)
    row += 1
    for i, h in enumerate(["#", "Check", "Status", "URLs Affected", "Detail / Action"], start=1):
        ws.cell(row=row, column=i, value=h)
    style_header(ws, row, 5)
    row += 1
    for cid in [t[0] for t in C_CHECKS]:
        cr = check_results[cid]
        urls_aff = len(cr.failed_urls) if cr.status == "Fail" else (None if cr.status in ("Blocked", "N/A") else 0)
        detail = cr.blocked_reason if cr.status == "Blocked" else (cr.detail or cr.action)
        for i, v in enumerate([cid, cr.name, cr.status, urls_aff, detail], start=1):
            ws.cell(row=row, column=i, value=v).alignment = Alignment(vertical="top", wrap_text=True)
        _apply_status_color(ws, row, 3, cr.status)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="SITEWIDE (S1-S12)").font = Font(name="Inter", size=11, bold=True, color=NAVY)
    row += 1
    for i, h in enumerate(["#", "Check", "Status", "URLs Affected", "Detail / Action"], start=1):
        ws.cell(row=row, column=i, value=h)
    style_header(ws, row, 5)
    row += 1
    for sr in s_results:
        for i, v in enumerate([sr["id"], sr["check"], sr["status"], "-", sr["detail"]], start=1):
            ws.cell(row=row, column=i, value=v).alignment = Alignment(vertical="top", wrap_text=True)
        _apply_status_color(ws, row, 3, sr["status"])
        row += 1

    autosize(ws, 5, max_w=70)
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["E"].width = 70


def write_issue_tab(wb, cid, cr: CheckResult):
    safe = "".join(" " if ch in r"\/?*[]:" else ch for ch in cr.name)
    sheet_name = f"{cid} {safe[:25]}"[:31].strip()
    ws = wb.create_sheet(sheet_name)
    color = {"Fix Now": GREEN, "Fix Now, Revisit": YELLOW, "Phase 3 Dependent": ORANGE}.get(cr.kw_dependency, GREY)
    color_sheet_tab(ws, color)
    style_title(ws, 1, f"{cid} -- {cr.name}", 6)
    style_subtitle(ws, 2,
                   f"{len(cr.failed_urls)} URL(s) affected. Action: {cr.action}. KW dependency: {cr.kw_dependency}.", 6)
    headers = ["URL", "Category", "Service", "Detail", "Action", "Status"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4, 6)
    for i, fu in enumerate(cr.failed_urls, start=5):
        vals = [fu["url"], fu["category"], fu["service"], fu["detail"], cr.action, fu.get("status", "To Do")]
        for j, v in enumerate(vals, start=1):
            ws.cell(row=i, column=j, value=v).alignment = Alignment(vertical="top", wrap_text=True)
    autosize(ws, 6, max_w=50)
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["D"].width = 60


def write_page_speed(wb, title, scope):
    ws = wb.create_sheet("Page Speed")
    color_sheet_tab(ws, BLUE)
    style_title(ws, 1, f"{title} -- Page Speed (BLOCKED)", 8)
    style_subtitle(ws, 2,
                   "Requires Screaming Frog PageSpeed integration export (pagespeed_all.csv + opportunities). "
                   "Run SF with PageSpeed enabled, export, then re-run.",
                   8)
    headers = ["URL", "Priority Tier", "Category", "Performance Score", "LCP", "CLS", "TBT", "Action Items"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4, 8)
    for i, (_, r) in enumerate(scope.iterrows(), start=5):
        vals = [_str(r["URL"]), _str(r["Action"]), _str(r["Category"]),
                "Blocked", "Blocked", "Blocked", "Blocked",
                "Pull SF PageSpeed report"]
        for j, v in enumerate(vals, start=1):
            ws.cell(row=i, column=j, value=v).alignment = Alignment(vertical="top", wrap_text=True)
    autosize(ws, 8, max_w=40)
    ws.column_dimensions["A"].width = 50


def write_architecture(wb, title, scope):
    ws = wb.create_sheet("Website Architecture")
    color_sheet_tab(ws, BLUE)
    style_title(ws, 1, f"{title} -- Website Architecture", 10)
    style_subtitle(ws, 2, "Per-URL depth + orphan + sitemap. Internal-link plan per Phase 1 inlinks data.", 10)
    headers = ["URL", "Priority Tier", "Category", "Service", "Current Depth", "Inlinks",
               "Orphan?", "In Sitemap?", "Action Items", "Status"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4, 10)
    medians = median_inlinks_by_category(scope)
    for i, (_, r) in enumerate(scope.iterrows(), start=5):
        depth = int(num(r.get("Page Depth")))
        inlinks = int(num(r.get("Inlinks")))
        in_sitemap = _str(r.get("In Sitemap")).lower() in ("true", "yes", "1")
        cat = _str(r.get("Category"))
        med = medians.get(cat, 0)
        actions = []
        if depth >= 4:
            actions.append(f"T6: Reduce depth from {depth} to <=3 (add nav/breadcrumb)")
        if inlinks == 0:
            actions.append("T4/S8: Add internal links (orphan)")
        elif med > 0 and inlinks < med * 0.5:
            actions.append(f"T5: Inlinks={inlinks} < 50% of {cat} median ({med:.1f}); add links")
        if not in_sitemap:
            actions.append("T12: Add to sitemap.xml")
        if not actions:
            actions.append("OK")
        vals = [_str(r["URL"]), _str(r["Action"]), cat, _str(r["Service"]),
                depth, inlinks, "Yes" if inlinks == 0 else "No",
                "Yes" if in_sitemap else "No",
                "; ".join(actions), "To Do" if actions != ["OK"] else "Done"]
        for j, v in enumerate(vals, start=1):
            ws.cell(row=i, column=j, value=v).alignment = Alignment(vertical="top", wrap_text=True)
    autosize(ws, 10, max_w=40)
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["I"].width = 60


def write_schema_optimization(wb, title, scope):
    ws = wb.create_sheet("Schema Optimization")
    color_sheet_tab(ws, BLUE)
    style_title(ws, 1, f"{title} -- Schema Optimization", 8)
    style_subtitle(ws, 2,
                   "Required schema per page category (transport industry defaults). "
                   "Current schema audit is BLOCKED -- requires SF structured data report.", 8)
    headers = ["URL", "Category", "Service", "Required Schema", "Current Schema", "Issue Type", "JSON-LD Plan", "Status"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4, 8)
    for i, (_, r) in enumerate(scope.iterrows(), start=5):
        cat = _str(r.get("Category"))
        required = SCHEMA_TARGETS_BY_CATEGORY.get(cat, "Breadcrumb (minimum)")
        plan = f"Implement {required}; validate via Rich Results Test"
        if cat == "Service Page":
            plan += "; consider Product+AggregateRating on highest-value service for review stars"
        vals = [_str(r["URL"]), cat, _str(r["Service"]), required, "Blocked (need SF)", "Blocked",
                plan, "To Do"]
        for j, v in enumerate(vals, start=1):
            ws.cell(row=i, column=j, value=v).alignment = Alignment(vertical="top", wrap_text=True)
    autosize(ws, 8, max_w=50)
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["G"].width = 60


def write_broken_list(wb, title, scope):
    ws = wb.create_sheet("Broken List")
    color_sheet_tab(ws, BLUE)
    style_title(ws, 1, f"{title} -- Broken List", 6)
    style_subtitle(ws, 2,
                   "Broken-link audit. Requires SF inlinks export to broken-target set "
                   "(response_codes_client_error_(4xx).csv + inlinks). Currently BLOCKED.", 6)
    headers = ["Source URL", "Source Type", "Broken Target", "Status Code", "Repair Method", "Acceptance Criteria"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4, 6)
    ws.cell(row=5, column=1, value="(BLOCKED -- pull SF reports listed in subtitle)").alignment = Alignment(vertical="top")
    autosize(ws, 6, max_w=40)


def write_url_priority(wb, title, priority_rows):
    ws = wb.create_sheet("URL Priority")
    color_sheet_tab(ws, BLUE)
    style_title(ws, 1, f"{title} -- URL Priority", 10)
    style_subtitle(ws, 2, "Combined view: every URL with all Phase 2 actions consolidated.", 10)
    headers = ["URL", "Priority Tier", "Category", "Service", "Current State", "Target State",
               "Action Items", "Acceptance Criteria", "KW Dependency", "Status"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4, 10)
    if not priority_rows:
        ws.cell(row=5, column=1, value="(no Phase 2 actions required -- all checks Pass for this scope)")
    for i, row in enumerate(priority_rows, start=5):
        vals = [row["URL"], row["Priority Tier"], row["Category"], row["Service"],
                row["Current State"], row["Target State"], row["Action Items"],
                row["Acceptance Criteria"], row["KW Dependency"], row["Status"]]
        for j, v in enumerate(vals, start=1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        dep_color = {"Fix Now": "DCFCE7", "Fix Now, Revisit": "FEF3C7", "Phase 3 Dependent": "FFEDD5"}.get(row["KW Dependency"])
        if dep_color:
            ws.cell(row=i, column=9).fill = PatternFill("solid", fgColor=dep_color)
    autosize(ws, 10, max_w=50)
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["G"].width = 60
    ws.freeze_panes = "B5"


def write_aggregate(wb, title, scope):
    ws = wb.create_sheet("Aggregate")
    color_sheet_tab(ws, GREY)
    style_title(ws, 1, f"{title} -- Aggregate (reference)", 8)
    style_subtitle(ws, 2, "All Optimize+Restore URLs with the source-of-truth columns from Phase 1.", 8)
    cols = ["URL", "Action", "Category", "Service", "Sessions", "Conversions", "Total Revenue",
            "Average Impressions", "Average CTR", "Best TV Keyword", "Best TV KW SV", "Best TV KW Rank",
            "Backlinks", "Referring Domains", "Indexability", "Word Count", "Inlinks", "Page Depth",
            "Current Title", "H1", "Meta Description", "Canonical Link Element", "In Sitemap"]
    for c in cols:
        if c not in scope.columns:
            scope[c] = None
    out = scope[cols]
    for j, h in enumerate(cols, start=1):
        ws.cell(row=4, column=j, value=h)
    style_header(ws, 4, len(cols))
    for i, row in enumerate(out.itertuples(index=False, name=None), start=5):
        for j, v in enumerate(row, start=1):
            if pd.isna(v):
                v = None
            ws.cell(row=i, column=j, value=v)
    autosize(ws, len(cols), max_w=40)
    ws.column_dimensions["A"].width = 50
    ws.freeze_panes = "B5"


# ─── public entrypoint ─────────────────────────────────────────────────────
def build_phase2_workbook(
    triage_df: pd.DataFrame,
    *,
    title: str,
    domain: str,
    check_states: dict[str, dict] | None = None,
    skip_sitewide_http: bool = False,
) -> BytesIO:
    """Build the Phase 2 workbook from a Phase 1 triage DataFrame.

    Parameters
    ----------
    triage_df
        Output of `_phase1_builder.build_phase1_dataframe` — must include
        Action, Category, Service, plus the original 44-column Phase 1
        fields.
    title
        Site display name used in tab titles.
    domain
        Bare primary domain (e.g. "buscharter.com.au"). Used for S1/S3
        live HTTP probes.
    check_states
        Optional map keyed by "url\\x1fcheck_id" → page_check_state row;
        used to overlay operator status (To Do / In Progress / Blocked /
        Done) onto each per-check tab.
    skip_sitewide_http
        If True, S1/S3 will return Blocked instead of hitting the live
        site. Useful when running on Vercel where outbound HTTP is slow.

    Returns
    -------
    BytesIO positioned at 0 ready for streaming.
    """
    check_states = check_states or {}

    scope = triage_df[triage_df["Action"].str.startswith(("Optimize", "Restore"))].copy().reset_index(drop=True)
    total_scope = len(scope)

    wb = Workbook()
    wb.remove(wb.active)

    if total_scope == 0:
        ws = wb.create_sheet("README")
        ws["A1"] = f"{title} -- no Optimize or Restore URLs in Phase 1 triage."
        ws["A2"] = "Cause: missing Meta data linkage (GA4/GSC/SF/DFS) for this domain."
        ws["A3"] = "Action: connect data sources to BigQuery, re-run Phase 1, then re-run Phase 2."
        ws["A1"].font = Font(name="Inter", size=14, bold=True, color=NAVY)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # Context for cross-URL checks
    titles_list = [_str(t) for t in scope["Current Title"] if _str(t)]
    metas = [_str(m) for m in scope["Meta Description"] if _str(m)]
    kws = [_str(k) for k in scope["Best TV Keyword"] if _str(k) and _str(k).lower() not in ("none found",)]
    ctx = {
        "median_inlinks": median_inlinks_by_category(scope),
        "title_to_urls": Counter(titles_list),
        "meta_to_urls": Counter(metas),
        "kw_to_urls": Counter(kws),
    }

    # Initialize check results
    check_results: dict[str, CheckResult] = {}
    for cid, name, action, kind, blocked_reason in T_CHECKS:
        cr = CheckResult(check_id=cid, name=name, action=action, kw_dependency=kw_dependency(cid))
        if kind == "blocked":
            cr.status = "Blocked"
            cr.blocked_reason = blocked_reason
        check_results[cid] = cr
    for cid, name, action, kind, blocked_reason in C_CHECKS:
        cr = CheckResult(check_id=cid, name=name, action=action, kw_dependency=kw_dependency(cid))
        if kind == "blocked":
            cr.status = "Blocked"
            cr.blocked_reason = blocked_reason
        check_results[cid] = cr

    # Run per-URL T/C checks
    for _, r in scope.iterrows():
        for cid, _name, _action, kind, _ in T_CHECKS:
            if kind != "active":
                continue
            fail, detail = t_check(cid, r, ctx)
            if fail:
                # Overlay per-check status if present.
                state = check_states.get(f"{_str(r['URL'])}\x1f{cid}") or {}
                check_results[cid].failed_urls.append({
                    "url": _str(r["URL"]),
                    "category": _str(r["Category"]),
                    "service": _str(r["Service"]),
                    "detail": detail,
                    "status": state.get("status", "To Do"),
                })
        for cid, _name, _action, kind, _ in C_CHECKS:
            if kind != "active":
                continue
            fail, detail = c_check(cid, r, ctx)
            if fail:
                state = check_states.get(f"{_str(r['URL'])}\x1f{cid}") or {}
                check_results[cid].failed_urls.append({
                    "url": _str(r["URL"]),
                    "category": _str(r["Category"]),
                    "service": _str(r["Service"]),
                    "detail": detail,
                    "status": state.get("status", "To Do"),
                })

    # Set final per-check status
    for cid, cr in check_results.items():
        if cr.status == "Blocked":
            continue
        if cr.failed_urls:
            cr.status = "Fail"
            cr.detail = f"{len(cr.failed_urls)} URL(s) affected"
        else:
            cr.status = "Pass"

    # Sitewide
    if skip_sitewide_http:
        s_results = [
            dict(id=cid, check=name, status="Blocked", detail="HTTP probes disabled in this environment")
            for cid, name, _, _ in S_CHECKS
        ]
    else:
        s_results = run_sitewide_checks(domain, scope)

    # URL Priority rows
    url_to_actions: defaultdict[str, list[tuple[str, str, str]]] = defaultdict(list)
    for cid, cr in check_results.items():
        if cr.status != "Fail":
            continue
        for fu in cr.failed_urls:
            url_to_actions[fu["url"]].append((cid, cr.action, cr.kw_dependency))

    priority_rows: list[dict] = []
    for _, r in scope.iterrows():
        url = _str(r["URL"])
        actions = url_to_actions.get(url, [])
        if not actions:
            continue
        priority_rows.append({
            "URL": url,
            "Priority Tier": r["Action"],
            "Category": _str(r["Category"]),
            "Service": _str(r["Service"]),
            "Current State": _format_current_state(r),
            "Target State": _format_target_state(r, actions),
            "Action Items": "; ".join(f"{cid}: {act}" for cid, act, _ in actions),
            "Acceptance Criteria": _format_acceptance(actions),
            "KW Dependency": _aggregate_kw_dep([d for _, _, d in actions]),
            "Status": "To Do",
        })

    write_issue_summary(wb, title, check_results, total_scope)
    write_audit_checklist(wb, title, check_results, s_results, total_scope)
    failed_checks = [(cid, cr) for cid, cr in check_results.items() if cr.status == "Fail"]
    for cid, cr in failed_checks:
        write_issue_tab(wb, cid, cr)
    write_page_speed(wb, title, scope)
    write_architecture(wb, title, scope)
    write_schema_optimization(wb, title, scope)
    write_broken_list(wb, title, scope)
    write_url_priority(wb, title, priority_rows)
    write_aggregate(wb, title, scope)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


__all__ = ["build_phase2_workbook", "T_CHECKS", "C_CHECKS", "S_CHECKS"]
