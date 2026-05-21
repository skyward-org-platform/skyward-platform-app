"""Phase 1 WQA workbook builder — ported from
`/Users/paulskirbe/agency/delivery/tna/build_phase1_wqa.py`.

This is a self-contained, in-memory variant of the same 12-tab workbook
that the CLI builder writes. It accepts an in-memory rows list (from
`/api/wqa/pages`) plus optional override maps (Supabase wqa_decision)
and emits a BytesIO buffer suitable for streaming as a Vercel response.

The public entrypoint is :func:`build_phase1_workbook`. Every other
function in this module is a near-verbatim copy of the original CLI
helpers, retained so the visual output is byte-identical to what the
delivery script produces.
"""
from __future__ import annotations

import re
from io import BytesIO
from typing import Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ─── COLUMN_RENAME: BQ snake_case → Title Case used by the builder ─────────
COLUMN_RENAME = {
    "page_path": "Page Path",
    "url": "URL",
    "is_primary_url": "Is Primary URL",
    "data_sources": "Data Sources",
    "in_sitemap": "In Sitemap",
    "best_tv_keyword": "Best TV Keyword",
    "best_tv_kw_sv": "Best TV KW SV",
    "best_tv_kw_rank": "Best TV KW Rank",
    "best_sv_keyword": "Best SV Keyword",
    "best_sv_kw_sv": "Best SV KW SV",
    "best_sv_kw_rank": "Best SV KW Rank",
    "type": "Type",
    "current_title": "Current Title",
    "meta_description": "Meta Description",
    "h1": "H1",
    "word_count": "Word Count",
    "link_score": "Link Score",
    "inlinks": "Inlinks",
    "outlinks": "Outlinks",
    "canonical_link_element": "Canonical Link Element",
    "status_code": "Status Code",
    "status": "Status",
    "indexability": "Indexability",
    "indexability_status": "Indexability Status",
    "page_depth": "Page Depth",
    "last_modified": "Last Modified",
    "sessions": "Sessions",
    "session_pct_change": "Session % Change",
    "losing_traffic": "Losing Traffic?",
    "average_session_duration": "Average Session Duration",
    "conversions": "Conversions",
    "conversion_rate_pct": "Conversion Rate (%)",
    "ecom_conversion_rate_pct": "Ecom Conversion Rate (%)",
    "total_revenue": "Total Revenue",
    "average_ctr": "Average CTR",
    "average_impressions": "Average Impressions",
    "backlinks": "Backlinks",
    "referring_domains": "Referring Domains",
    "dofollow": "DoFollow",
    "nofollow": "NoFollow",
}


# ─── helpers ────────────────────────────────────────────────────────────────
def num(v):
    if v is None:
        return 0
    if isinstance(v, str):
        try:
            return float(v)
        except ValueError:
            return 0
    if pd.isna(v):
        return 0
    try:
        return float(v)
    except Exception:
        return 0


def url_path(url: str) -> str:
    return re.sub(r"^https?://[^/]+", "", url or "")


def first_segment(url: str) -> str:
    p = url_path(url).lstrip("/").rstrip("/")
    return p.split("/")[0] if p else ""


AU_STATES = {
    "nsw", "new-south-wales", "victoria", "vic", "queensland", "qld",
    "western-australia", "wa", "south-australia", "sa", "tasmania", "tas",
    "northern-territory", "nt", "act", "australian-capital-territory",
}
NZ_REGIONS = {
    "auckland", "wellington", "christchurch", "hamilton", "tauranga", "dunedin",
    "north-island", "south-island", "queenstown", "rotorua", "napier",
}
AU_NZ_CITIES = {
    "sydney", "melbourne", "brisbane", "perth", "adelaide", "gold-coast",
    "newcastle", "canberra", "hobart", "darwin", "wollongong", "geelong",
    "cairns", "townsville", "ballarat", "bendigo", "launceston", "mackay",
    "rockhampton", "toowoomba", "sunshine-coast",
    "auckland", "wellington", "christchurch", "hamilton", "tauranga",
    "dunedin", "queenstown", "rotorua", "napier", "palmerston-north",
    "nelson", "new-plymouth", "whangarei",
}

UTILITY_PATHS = {
    "about", "about-us", "contact", "contact-us", "terms", "terms-of-use",
    "terms-and-conditions", "privacy", "privacy-policy", "sitemap",
    "faq", "faqs", "careers", "team", "our-team", "history", "press",
    "media", "media-kit", "blog-rss", "feed", "thank-you", "thanks",
    "subscribe", "unsubscribe", "newsletter",
}

QUOTE_KEYWORDS = ("quote", "enquire", "enquiry", "book", "booking", "rfq", "contact")


def classify_category(url: str, services: tuple[str, ...]) -> str:
    if not url:
        return "Non-Analyzable"
    u = url.lower()
    path = url_path(u)
    p = path.lstrip("/").rstrip("/")

    if "?" in u and any(t in u for t in ("utm_", "gclid", "fbclid", "_gl=", "?ver=", "msclkid")):
        return "Parameter Variant"
    if "#" in url and url.split("#", 1)[1]:
        return "Fragment"
    if any(x in u for x in (
        "/wp-content", "/wp-includes", "/wp-json", "/cdn-cgi", "/_next/",
        "/feed", "/xmlrpc.php", "/sitemap.xml",
    )):
        return "Non-Analyzable"
    if u.endswith((".jpg", ".jpeg", ".png", ".gif", ".svg", ".webp", ".pdf",
                   ".css", ".js", ".xml", ".ico", ".txt")):
        return "Non-Analyzable"
    if u.startswith(("tel:", "mailto:", "sms:", "javascript:")):
        return "Non-Analyzable"

    if path in ("", "/"):
        return "Homepage"

    seg = first_segment(u)
    segments = [s for s in p.split("/") if s]
    last = segments[-1] if segments else ""

    if seg in UTILITY_PATHS or last in UTILITY_PATHS:
        return "Utility"

    if any(k in p for k in QUOTE_KEYWORDS):
        return "Quote/Contact"

    if seg in ("blog", "news", "articles", "resources"):
        if not segments[1:]:
            return "Blog Hub"
        return "Blog Post"
    if seg in ("category", "tag", "author"):
        return "Blog Category"

    if seg in ("fleet", "vehicles", "vehicle", "our-fleet"):
        return "Fleet/Product Page"

    if seg in AU_STATES or last in AU_STATES:
        return "Location Page"
    if seg in NZ_REGIONS or last in NZ_REGIONS:
        return "Location Page"
    if seg in AU_NZ_CITIES or last in AU_NZ_CITIES:
        return "Location Page"

    seg_compact = seg.replace("_", "-")
    last_compact = last.replace("_", "-")
    if any(svc in seg_compact for svc in services) or any(svc in last_compact for svc in services):
        return "Service Page"
    if "tour" in seg or "tour" in last or "transfer" in seg or "transfer" in last:
        return "Service Page"

    return "Other"


def classify_location(url: str) -> tuple[str, str]:
    if not url:
        return ("None found", "None found")
    path = url_path(url.lower())
    p = path.lstrip("/").rstrip("/")
    segments = [s for s in p.split("/") if s]
    for s in segments[::-1]:
        sl = s.replace("_", "-")
        if sl in AU_STATES:
            return ("State", sl.replace("-", " ").title())
        if sl in NZ_REGIONS:
            return ("Region", sl.replace("-", " ").title())
        if sl in AU_NZ_CITIES:
            return ("City", sl.replace("-", " ").title())
    return ("None found", "None found")


def classify_service(url: str, category: str, services: tuple[str, ...]) -> str:
    if not url:
        return "Non-Analyzable"
    u = url.lower()
    if category == "Homepage":
        return "General (Homepage)"
    if category == "Quote/Contact":
        return "Quote / Contact"
    if category in ("Blog Hub", "Blog Post", "Blog Category"):
        return "Content / Blog"
    if category == "Utility":
        return "Utility"
    if category == "Fleet/Product Page":
        return "Fleet"
    if category == "Location Page":
        return "Bus Hire by Location"
    if category == "Service Page":
        if "party" in u or "wedding" in u or "hens" in u or "bucks" in u:
            return "Party / Event"
        if "wine" in u or "tour" in u:
            return "Tours"
        if "school" in u:
            return "School Transport"
        if "corporate" in u:
            return "Corporate Transport"
        if "wedding" in u:
            return "Weddings"
        if "airport" in u or "transfer" in u:
            return "Airport / Transfers"
        if "mini" in u:
            return "Minibus Hire"
        return "Bus / Coach Hire"
    if category == "Non-Analyzable":
        return "Non-Analyzable"
    if category == "Parameter Variant":
        return "Non-Analyzable"
    if category == "Fragment":
        return "Non-Analyzable"
    return "Other"


def is_primary(status, indexability) -> int:
    if pd.isna(status):
        return 0
    try:
        s = int(float(status))
    except (TypeError, ValueError):
        return 0
    if s == 200 and indexability == "Indexable":
        return 1
    return 0


def signal_string(r: pd.Series) -> str:
    parts = []
    sess = num(r.get("Sessions"))
    imp = num(r.get("Average Impressions"))
    rd = num(r.get("Referring Domains"))
    bl = num(r.get("Backlinks"))
    conv = num(r.get("Conversions"))
    if sess > 0:
        parts.append(f"GA4: {int(sess)} sessions")
    if imp > 0:
        parts.append(f"GSC: {int(imp)} impressions")
    if rd > 0:
        parts.append(f"Ahrefs: {int(rd)} ref domains ({int(bl)} backlinks)")
    if conv > 0:
        parts.append(f"GA4: {int(conv)} conversions")
    kw = r.get("Best TV Keyword")
    if pd.notna(kw) and kw not in (None, "None found", ""):
        sv = int(num(r.get("Best TV KW SV")))
        rk = int(num(r.get("Best TV KW Rank")))
        parts.append(f'DFS: "{kw}" (rank {rk}, SV {sv})')
    return " | ".join(parts) if parts else "Zero sessions, zero impressions, zero ref domains, no rankings"


def has_signals(r: pd.Series) -> bool:
    return (
        num(r.get("Sessions")) > 0
        or num(r.get("Average Impressions")) > 0
        or num(r.get("Referring Domains")) > 0
        or num(r.get("Backlinks")) > 0
    )


def is_ranking_page1(r):
    rk = num(r.get("Best TV KW Rank"))
    sv = num(r.get("Best TV KW SV"))
    return 1 <= rk <= 10 and sv >= 50


def is_striking_distance(r):
    rk = num(r.get("Best TV KW Rank"))
    sv = num(r.get("Best TV KW SV"))
    return 11 <= rk <= 20 and sv >= 100


def assign_action(r: pd.Series) -> tuple[str, str]:
    url = r.get("URL") or ""
    status = r.get("Status Code")
    idx = r.get("Indexability")
    idx_status = r.get("Indexability Status")
    cat = r["Category"]
    sess = num(r.get("Sessions"))
    conv = num(r.get("Conversions"))
    rd = num(r.get("Referring Domains"))
    bl = num(r.get("Backlinks"))
    imp = num(r.get("Average Impressions"))
    sig = signal_string(r)

    if cat == "Fragment":
        return ("No Action (fragment)", f"Fragment URL. Anchor on parent. {sig}")
    if cat == "Non-Analyzable":
        return ("No Action (system URL)", f"System URL. Not a rankable page. {sig}")
    if cat == "Parameter Variant":
        canon = r.get("Canonical Link Element")
        if pd.notna(canon) and canon:
            return ("No Action (parameter variant)", f"Parameter variant. Canonical to {canon}. {sig}")
        return ("No Action (parameter variant)", f"Parameter variant (UTM, gclid, fbclid). Authority consolidates to canonical. {sig}")

    try:
        sc = int(float(status)) if pd.notna(status) else None
    except (TypeError, ValueError):
        sc = None

    if sc == 301:
        if rd >= 3 or sess >= 30:
            return ("Redirect (already handled, audit chain)",
                    f"301 with material authority. Verify chain length and destination. {sig}")
        return ("Redirect (already handled)", f"Already 301. Working correctly. {sig}")
    if sc == 302:
        return ("Redirect (convert 302 to 301)", f"302 (temporary) should be a permanent 301. {sig}")

    if sc == 404 or (sc is not None and 400 <= sc < 500 and sc != 403):
        rk = num(r.get("Best TV KW Rank"))
        sv = num(r.get("Best TV KW SV"))
        if sess >= 30 or (1 <= rk <= 20 and sv >= 100):
            return ("Restore (200)",
                    f"404 with significant value (traffic or ranking keyword). Restore page to 200. {sig}")
        if rd >= 3 or bl >= 5:
            return ("Redirect (broken with authority)",
                    f"404 with link equity at risk. 301 to closest live page. {sig}")
        if has_signals(r):
            return ("Redirect (broken with some value)",
                    f"404 with minor value. 301 to closest live page. {sig}")
        return ("No Action (no value)", f"404 with zero signals. Leave as 404. {sig}")

    if sc == 403:
        if cat == "Non-Analyzable":
            return ("No Action (system URL)", f"403 system path. {sig}")
        if has_signals(r):
            return ("Review (unexpected 403)", f"403 with material signals. Investigate. {sig}")
        return ("No Action (no value)", f"403 with zero signals. {sig}")

    if sc is not None and 500 <= sc < 600:
        if has_signals(r):
            return ("Restore (fix server error)",
                    f"5xx server error on page with signals. Investigate platform issue. {sig}")
        return ("Investigate (5xx)", f"Server error. Determine if page should exist. {sig}")

    if idx_status == "Canonicalised":
        canon = r.get("Canonical Link Element") or ""
        return ("No Action (already handled)", f"Canonicalised to {canon}. {sig}")
    if idx_status == "noindex":
        if sess >= 30 or rd > 0 or imp >= 10:
            return ("Review (intentional noindex with traffic)",
                    f"Marked noindex but has signals. Confirm intent. {sig}")
        return ("No Action (intentional noindex)", f"Intentionally noindexed. {sig}")

    if sc == 200 and idx == "Non-Indexable":
        return ("Review (unexpected non-indexable)",
                f"Status 200 but Non-Indexable ({idx_status}). Review. {sig}")

    if sc == 200 and idx == "Indexable":
        if cat == "Homepage":
            return ("Optimize (revenue-critical)", f"Core homepage. Highest priority. {sig}")
        if cat == "Quote/Contact" or conv >= 50:
            return ("Optimize (revenue-critical)", f"Conversion / quote page. Revenue-critical. {sig}")
        if conv >= 20 or sess >= 500:
            return ("Optimize (revenue-critical)", f"High-traffic / high-conversion page. {sig}")
        if is_ranking_page1(r):
            return ("Optimize (page 1 - protect/improve)", f"Ranking page 1. Protect + improve. {sig}")
        if is_striking_distance(r):
            return ("Optimize (striking distance)", f"Ranking page 2. Push to page 1. {sig}")
        if has_signals(r):
            if cat == "Utility":
                return ("Optimize (utility - light touch)", f"Utility page with signals. Metadata only. {sig}")
            return ("Optimize (has visibility)", f"Has traffic, search visibility, or authority. {sig}")
        wc = num(r.get("Word Count"))
        inlinks = num(r.get("Inlinks"))
        if cat == "Utility" and wc >= 300:
            return ("Optimize (utility - light touch)", f"Utility page. Light-touch metadata work. {sig}")
        if cat in ("Blog Hub", "Fleet/Product Page") and inlinks >= 3:
            return ("Optimize (utility - light touch)", f"Hub page. Architectural role. {sig}")
        if cat == "Blog Post" and wc < 300:
            return ("Remove (thin content)", f"Thin blog post (<300 words) with zero signals. {sig}")
        if cat == "Location Page" and not has_signals(r):
            return ("Consolidate (no demand)", f"Location page with no search demand. Merge into parent. {sig}")
        if inlinks >= 5:
            return ("Evaluate (linked, no external signals)",
                    f"{int(inlinks)} internal links but no external signals. {sig}")
        if inlinks >= 2:
            return ("Evaluate (some internal links)",
                    f"{int(inlinks)} internal links, no external signals. {sig}")
        if cat == "Utility":
            return ("Remove (no signals)", f"Utility page with zero signals and minimal content. {sig}")
        return ("Remove (no signals)", f"No value signals from any source. {sig}")

    if sc is None:
        if cat == "Non-Analyzable":
            return ("No Action (system URL)", f"System URL surfaced via GA4 but not crawled. {sig}")
        if cat == "Parameter Variant":
            return ("No Action (parameter variant)", f"Parameter variant in GA4. Traffic belongs to clean URL. {sig}")
        if (url or "").startswith("http://") and rd > 0:
            return ("Redirect (http to https)", f"http variant with backlinks. Should 301 to https. {sig}")
        if sess >= 30:
            return ("Redirect (broken variant with value)", f"GA4 ghost URL with traffic. {sig}")
        if rd >= 3:
            return ("Redirect (broken variant with authority)", f"Backlinks-only ghost URL. {sig}")
        return ("No Action (no value)", f"GA4/Backlinks ghost URL not in current crawl. {sig}")

    return ("Review (unexpected status)", f"Status {status} / Indexability {idx}. Manual review. {sig}")


def health_zone(action: str) -> str:
    if action.startswith("Optimize"):
        return "Green"
    if action.startswith("Restore"):
        return "Yellow"
    if action.startswith("Review"):
        return "Yellow"
    if action.startswith("Redirect"):
        return "Orange"
    if action.startswith("Remove"):
        return "Red"
    if action.startswith("Consolidate"):
        return "Orange"
    if action.startswith("Evaluate"):
        return "Yellow"
    if action.startswith("Investigate"):
        return "Yellow"
    return "Gray"


# ─── styling ────────────────────────────────────────────────────────────────
NAVY = "0F2A47"
SLATE = "475569"
WHITE = "FFFFFF"
THIN = Side(border_style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name="Inter", size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
        cell.border = BORDER


def style_title(ws, row, text, ncols=8):
    ws.cell(row=row, column=1, value=text)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1)
    cell.font = Font(name="Inter", size=14, bold=True, color=NAVY)


def style_subtitle(ws, row, text, ncols=8):
    ws.cell(row=row, column=1, value=text)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1)
    cell.font = Font(name="Inter", size=10, italic=True, color=SLATE)
    cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)


def autosize(ws, max_col, max_w=80):
    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        max_len = 8
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            v = row[0]
            if v is None:
                continue
            l = len(str(v))
            if l > max_len:
                max_len = l
        ws.column_dimensions[letter].width = min(max_len + 2, max_w)


# ─── tabs ───────────────────────────────────────────────────────────────────
def write_action_legend(wb):
    ws = wb.create_sheet("Action Legend")
    style_title(ws, 1, "Action Legend", 3)
    style_subtitle(ws, 2, "Every URL in URL Triage gets exactly one action.", 3)
    rows = [
        ("Action", "What It Means", "What Happens Next"),
        ("Optimize", "URL stays. At least one positive signal.",
         "Enters Phase 2/3 optimization pipeline."),
        ("Redirect", "Needs a 301 to a better URL.",
         "Listed in Redirect Map. Developer executes 301."),
        ("Restore", "Broken (404/5xx) but should exist (active service, has rankings/traffic).",
         "Listed in Restore URLs. Content team specs, dev restores."),
        ("Remove", "No value signals. Noindex or delete.",
         "Listed in Removal List. Skyward applies noindex."),
        ("Consolidate", "Duplicates a parent in intent. Merge content and 301.",
         "Listed in Canonicalization Map."),
        ("Evaluate", "Has internal links but no external signals. Needs human judgment.",
         "Reviewed at Checkpoint 1."),
        ("Review", "Some signals but action not obvious.",
         "Flagged for review. Logic explains."),
        ("Investigate", "Data conflict (5xx with signals, primary that redirects).",
         "Manual investigation."),
        ("No Action", "Duplicate variant, system URL, fragment, intentionally excluded.",
         "Accounted for, no work."),
    ]
    for i, (a, b, c) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=a)
        ws.cell(row=i, column=2, value=b)
        ws.cell(row=i, column=3, value=c)
        if i == 4:
            style_header(ws, i, 3)
        else:
            for c_ in range(1, 4):
                cell = ws.cell(row=i, column=c_)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.font = Font(name="Inter", size=10)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 60
    for r in range(4, 4 + len(rows)):
        ws.row_dimensions[r].height = 36


def write_action_plan(wb, rows):
    ws = wb.create_sheet("Action Plan")
    style_title(ws, 1, "Action Plan", 6)
    style_subtitle(ws, 2, "Phase 1 to-dos, prioritized.", 6)
    headers = ["#", "Action Item", "URLs", "Severity", "Owner", "Cross-Reference"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4, 6)
    for i, row in enumerate(rows, start=5):
        for j, v in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=v).alignment = Alignment(vertical="top", wrap_text=True)
    autosize(ws, 6)
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["F"].width = 50


def write_url_triage(wb, df: pd.DataFrame):
    ws = wb.create_sheet("URL Triage")
    rename = {
        "Average Impressions": "Avg Impressions",
        "Average CTR": "Avg CTR",
        "Best TV KW SV": "Best TV SV",
        "Best TV KW Rank": "Best TV Rank",
        "Referring Domains": "Ref Domains",
        "Current Title": "Title",
        "Canonical Link Element": "Canonical",
    }
    out = df.rename(columns=rename).copy()
    out["Health Zone"] = out["Action"].apply(health_zone)
    cols = ["URL", "Action", "Logic", "Status Code", "Category", "Location Type", "Location",
            "Service", "Primary?", "Sessions", "Conversions", "Total Revenue",
            "Avg Impressions", "Avg CTR", "Best TV Keyword", "Best TV SV", "Best TV Rank",
            "Backlinks", "Ref Domains", "Indexability", "Indexability Status",
            "Word Count", "Inlinks", "Page Depth", "Title", "H1", "Canonical",
            "In Sitemap", "Health Zone", "Data Sources"]
    for c in cols:
        if c not in out.columns:
            out[c] = None
    out = out[cols]
    for j, h in enumerate(cols, start=1):
        ws.cell(row=1, column=j, value=h)
    style_header(ws, 1, len(cols))
    ws.freeze_panes = "B2"
    for i, row in enumerate(out.itertuples(index=False, name=None), start=2):
        for j, v in enumerate(row, start=1):
            if pd.isna(v):
                v = None
            ws.cell(row=i, column=j, value=v)
    fills = {"Optimize": "DCFCE7", "Restore": "FEF3C7", "Redirect": "FFEDD5",
             "Remove": "FECACA", "Review": "FEF9C3", "No Action": "F1F5F9",
             "Consolidate": "E9D5FF", "Evaluate": "DBEAFE", "Investigate": "FCE7F3"}
    for i in range(2, len(out) + 2):
        action = ws.cell(row=i, column=2).value or ""
        for prefix, color in fills.items():
            if action.startswith(prefix):
                ws.cell(row=i, column=2).fill = PatternFill("solid", fgColor=color)
                break
    autosize(ws, len(cols), max_w=50)
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 70


def write_funnel_summary(wb, df, total, title):
    ws = wb.create_sheet("Funnel Summary")
    style_title(ws, 1, f"{title} -- URL Triage Funnel", 3)
    style_subtitle(ws, 2, "Every URL discovered. Nothing hidden.", 3)
    ws.cell(row=4, column=1, value="TOTAL URLs").font = Font(name="Inter", bold=True, size=11, color=NAVY)
    ws.cell(row=4, column=2, value=total).font = Font(name="Inter", bold=True, size=11, color=NAVY)
    headers = ["Action", "URLs", "%"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=6, column=i, value=h)
    style_header(ws, 6, 3)
    counts = df["Action"].value_counts()
    groups = {
        "OPTIMIZE": [a for a in counts.index if a.startswith("Optimize")],
        "RESTORE": [a for a in counts.index if a.startswith("Restore")],
        "REDIRECT": [a for a in counts.index if a.startswith("Redirect")],
        "CONSOLIDATE": [a for a in counts.index if a.startswith("Consolidate")],
        "REMOVE": [a for a in counts.index if a.startswith("Remove")],
        "REVIEW": [a for a in counts.index if a.startswith("Review")],
        "EVALUATE": [a for a in counts.index if a.startswith("Evaluate")],
        "INVESTIGATE": [a for a in counts.index if a.startswith("Investigate")],
        "NON-ADDRESSABLE": [a for a in counts.index if a.startswith("No Action")],
    }
    r = 7
    for parent, subs in groups.items():
        if not subs:
            continue
        total_grp = sum(counts[s] for s in subs)
        ws.cell(row=r, column=1, value=parent).font = Font(name="Inter", bold=True, color=NAVY)
        ws.cell(row=r, column=2, value=int(total_grp)).font = Font(name="Inter", bold=True, color=NAVY)
        pct = f"{total_grp/total*100:.1f}%" if total else "0.0%"
        ws.cell(row=r, column=3, value=pct).font = Font(name="Inter", bold=True, color=NAVY)
        r += 1
        for s in subs:
            sub_label = s.split("(", 1)[1].rstrip(")") if "(" in s else s
            ws.cell(row=r, column=1, value=f"  {sub_label}")
            ws.cell(row=r, column=2, value=int(counts[s]))
            ws.cell(row=r, column=3, value=f"{counts[s]/total*100:.1f}%" if total else "0.0%")
            r += 1
        r += 1
    autosize(ws, 3)
    ws.column_dimensions["A"].width = 50


def write_service_summary(wb, df):
    ws = wb.create_sheet("Service Summary")
    style_title(ws, 1, "Service x Action", 6)
    style_subtitle(ws, 2, "How URLs and traffic distribute across services.", 6)
    headers = ["Service", "Total URLs", "Optimize", "Redirect", "Remove", "Sessions"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4, 6)
    g = df.groupby("Service")
    rows = []
    for svc, sub in g:
        opt = sub["Action"].str.startswith("Optimize").sum()
        red = sub["Action"].str.startswith("Redirect").sum()
        rem = sub["Action"].str.startswith("Remove").sum()
        sess = sub["Sessions"].apply(num).sum()
        rows.append((svc, len(sub), int(opt), int(red), int(rem), int(sess)))
    rows.sort(key=lambda x: -x[5])
    for i, row in enumerate(rows, start=5):
        for j, v in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=v)
    autosize(ws, 6)


def guess_destination(url: str, primary_host: str) -> str:
    if not url:
        return f"{primary_host}/"
    u = url.lower()
    seg = first_segment(u)
    if seg in ("blog", "news", "articles") and not url_path(u).rstrip("/").lstrip("/").startswith(seg + "/"):
        return f"{primary_host}/blog/"
    if seg in ("category", "tag", "author"):
        return f"{primary_host}/blog/"
    if any(k in u for k in QUOTE_KEYWORDS):
        return f"{primary_host}/contact/"
    if u.startswith("http://"):
        return "https://" + u[len("http://"):]
    last = url.rstrip("/").split("/")[-1].lower()
    if last in AU_NZ_CITIES:
        return f"{primary_host}/locations/"
    return f"{primary_host}/"


def write_redirect_map(wb, df, primary_host, title):
    ws = wb.create_sheet("Redirect Map")
    style_title(ws, 1, f"{title} -- Redirect Map", 9)
    style_subtitle(ws, 2, "Every URL with a Redirect action.", 9)
    headers = ["#", "Source URL", "Triage Action", "Destination URL", "Type", "Priority",
               "Sessions", "Ref Domains", "Triage Logic"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4, 9)
    redirects = df[df["Action"].str.startswith("Redirect")].copy()
    redirects["__sess"] = redirects["Sessions"].apply(num)
    redirects = redirects.sort_values("__sess", ascending=False)
    if len(redirects) == 0:
        ws.cell(row=5, column=1, value="(none)")
    for i, (_, r) in enumerate(redirects.iterrows(), start=5):
        url = r["URL"]
        action = r["Action"]
        sess = int(num(r.get("Sessions")))
        rd = int(num(r.get("Referring Domains")))
        # Use operator-set target URL if available, else SOP guess.
        target_override = r.get("Target URL")
        dest = (
            target_override
            if (pd.notna(target_override) and target_override)
            else guess_destination(url, primary_host)
        )
        priority = "High" if sess >= 50 or rd >= 5 else "Medium" if sess >= 10 or rd >= 1 else "Low"
        rtype = "301 (existing)" if "already handled" in action else "Standard 301"
        for j, v in enumerate([i - 4, url, action, dest, rtype, priority, sess, rd, r["Logic"]], start=1):
            ws.cell(row=i, column=j, value=v).alignment = Alignment(vertical="top", wrap_text=True)
    autosize(ws, 9, max_w=50)
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["I"].width = 60


def write_canonicalization_map(wb, df, primary_host, title):
    ws = wb.create_sheet("Canonicalization Map")
    style_title(ws, 1, f"{title} -- Canonicalization Map", 6)
    style_subtitle(ws, 2, "Consolidate decisions and non-primary 200s with explicit canonicals.", 6)
    headers = ["#", "Canonical Keeper", "Absorbed Duplicate", "Category", "Type", "Reason"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4, 6)
    rows = []
    for _, r in df.iterrows():
        if r["Action"].startswith("Consolidate"):
            keeper = guess_destination(r["URL"], primary_host)
            rows.append((keeper, r["URL"], r["Category"], r["Action"], r["Logic"]))
        elif r.get("Indexability Status") == "Canonicalised":
            canon = r.get("Canonical Link Element") or ""
            rows.append((canon, r["URL"], r["Category"], "Canonicalised (already)", r["Logic"]))
    if not rows:
        ws.cell(row=5, column=1, value="(no consolidation needed)")
    for i, row in enumerate(rows, start=5):
        ws.cell(row=i, column=1, value=i - 4)
        for j, v in enumerate(row, start=2):
            ws.cell(row=i, column=j, value=v).alignment = Alignment(vertical="top", wrap_text=True)
    autosize(ws, 6, max_w=50)
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["F"].width = 60


def write_canonical_audit(wb, df, title):
    ws = wb.create_sheet("Canonical Audit")
    style_title(ws, 1, f"{title} -- Canonical Audit", 5)
    style_subtitle(ws, 2, "Every Optimize URL: current canonical vs correct canonical.", 5)
    headers = ["URL", "Current Canonical", "Correct Canonical", "Issue Type", "Notes"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4, 5)
    optimize = df[df["Action"].str.startswith("Optimize")]
    if len(optimize) == 0:
        ws.cell(row=5, column=1, value="(no optimize URLs)")
    for i, (_, r) in enumerate(optimize.iterrows(), start=5):
        url = r["URL"]
        canon = r.get("Canonical Link Element") or ""
        if pd.isna(canon):
            canon = ""
        correct = url
        if not canon:
            issue, notes = "Missing", "No canonical tag detected. Add self-canonical."
        elif canon == url:
            issue, notes = "OK", "Self-canonical correct."
        else:
            issue, notes = "Mismatch", f"Current canonical ({canon}) does not match URL. Verify."
        for j, v in enumerate([url, canon, correct, issue, notes], start=1):
            ws.cell(row=i, column=j, value=v).alignment = Alignment(vertical="top", wrap_text=True)
    autosize(ws, 5, max_w=50)
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 50


def write_removal_list(wb, df, title):
    ws = wb.create_sheet("Removal List")
    style_title(ws, 1, f"{title} -- Removal List", 7)
    style_subtitle(ws, 2, "URLs to noindex/delete or leave as 404.", 7)
    headers = ["URL", "Current Status", "Recommended Action", "Category", "Word Count", "Sessions", "Reason"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4, 7)
    removes = df[df["Action"].str.startswith("Remove") | (df["Action"] == "No Action (no value)")].copy()
    if len(removes) == 0:
        ws.cell(row=5, column=1, value="(none)")
    for i, (_, r) in enumerate(removes.iterrows(), start=5):
        url = r["URL"]
        status = r.get("Status Code")
        try:
            sc = int(float(status)) if pd.notna(status) else None
        except (TypeError, ValueError):
            sc = None
        if sc == 200:
            rec = "Noindex + remove from nav"
        elif sc and 400 <= sc < 600:
            rec = "Leave as 404" if sc == 404 else "Investigate"
        else:
            rec = "Leave (no action)"
        for j, v in enumerate([url, sc if sc else "", rec, r["Category"],
                              int(num(r.get("Word Count"))), int(num(r.get("Sessions"))), r["Logic"]], start=1):
            ws.cell(row=i, column=j, value=v).alignment = Alignment(vertical="top", wrap_text=True)
    autosize(ws, 7, max_w=50)
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["G"].width = 60


def write_implementation_checklist(wb, counts, title):
    ws = wb.create_sheet("Implementation Checklist")
    style_title(ws, 1, f"{title} -- Implementation Checklist", 7)
    style_subtitle(ws, 2, "Phase 1 actions, sequenced.", 7)
    headers = ["#", "Step", "Owner", "URLs", "Severity", "Depends On", "Done"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4, 7)
    rows = [
        (1, f"Execute Redirect Map ({counts.get('redirect',0)} redirects, prioritized)", "Client Dev",
         counts.get("redirect", 0), "High", "-", ""),
        (2, f"Restore {counts.get('restore',0)} broken pages with material value", "Client Dev + Skyward content",
         counts.get("restore", 0), "High" if counts.get("restore",0) > 0 else "Low", "-", ""),
        (3, f"Apply noindex / delete on {counts.get('remove',0)} zero-value URLs", "Skyward",
         counts.get("remove", 0), "Medium", "Redirects done", ""),
        (4, f"Consolidate {counts.get('consolidate',0)} duplicate template pages", "Skyward + Client Dev",
         counts.get("consolidate", 0), "Medium", "-", ""),
        (5, f"Resolve {counts.get('review',0)} Review + {counts.get('evaluate',0)} Evaluate items",
         "Skyward + Client", counts.get("review", 0) + counts.get("evaluate", 0), "Medium", "-", ""),
        (6, "Audit canonicals on Optimize URLs", "Client Dev", counts.get("optimize", 0), "Medium",
         "Optimize set finalized", ""),
        (7, "Re-crawl with Screaming Frog after fixes", "Skyward", "", "Low", "Steps 1-6 complete", ""),
        (8, f"Hand {counts.get('optimize',0) + counts.get('restore',0)} URLs to Phase 2 + Phase 3",
         "Skyward", counts.get("optimize", 0) + counts.get("restore", 0), "High", "Triage signed off", ""),
    ]
    for i, row in enumerate(rows, start=5):
        for j, v in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=v).alignment = Alignment(vertical="top", wrap_text=True)
    autosize(ws, 7)
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["F"].width = 36


def write_url_optimization(wb, df, title):
    ws = wb.create_sheet("URL Optimization")
    style_title(ws, 1, f"{title} -- URL Optimization (Phase 2 Handoff)", 24)
    style_subtitle(ws, 2,
                   "Optimize + Restore URLs. Phase 2 layers technical, Phase 3 layers keyword cluster.", 24)
    cols = ["URL", "Priority Tier", "Triage Action", "Category", "Location Type", "Location",
            "Service", "Best TV Keyword", "Best TV SV", "Best TV Rank", "Sessions",
            "Conversions", "Total Revenue", "Avg Impressions", "Avg CTR",
            "Backlinks", "Ref Domains", "Word Count", "Inlinks", "Page Depth",
            "Title", "H1", "Canonical", "In Sitemap"]
    for i, h in enumerate(cols, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4, len(cols))
    pri_map = {
        "Optimize (revenue-critical)": "1. Revenue-Critical",
        "Optimize (page 1 - protect/improve)": "2. Page 1 Protect",
        "Optimize (striking distance)": "3. Striking Distance",
        "Optimize (has visibility)": "4. Has Visibility",
        "Optimize (utility - light touch)": "5. Utility",
        "Restore (200)": "1. Restore (high value)",
        "Restore (fix server error)": "1. Restore (5xx)",
    }
    sub = df[df["Action"].str.startswith(("Optimize", "Restore"))].copy()
    sub["Priority Tier"] = sub["Action"].map(pri_map).fillna("6. Other")
    sub["__sess"] = sub["Sessions"].apply(num)
    sub = sub.sort_values(["Priority Tier", "__sess"], ascending=[True, False])
    rename = {"Average Impressions": "Avg Impressions", "Average CTR": "Avg CTR",
              "Best TV KW SV": "Best TV SV", "Best TV KW Rank": "Best TV Rank",
              "Referring Domains": "Ref Domains", "Current Title": "Title",
              "Canonical Link Element": "Canonical"}
    sub = sub.rename(columns=rename)
    sub["Triage Action"] = sub["Action"]
    for c in cols:
        if c not in sub.columns:
            sub[c] = None
    out = sub[cols]
    for i, row in enumerate(out.itertuples(index=False, name=None), start=5):
        for j, v in enumerate(row, start=1):
            if pd.isna(v):
                v = None
            ws.cell(row=i, column=j, value=v)
    autosize(ws, len(cols), max_w=40)
    ws.column_dimensions["A"].width = 50
    ws.freeze_panes = "B5"


def write_restore_urls(wb, df, title):
    ws = wb.create_sheet("Restore URLs")
    style_title(ws, 1, f"{title} -- Restore URLs", 12)
    style_subtitle(ws, 2, "Pages to restore. Phase 4 fills the content spec.", 12)
    headers = ["URL", "Current Status", "Sessions", "Ref Domains", "Best Keyword",
               "Rank", "Search Volume", "Why Restore", "Recommended Action",
               "Target H1", "Target Title", "Target Meta"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4, 12)
    sub = df[df["Action"].str.startswith("Restore")]
    if len(sub) == 0:
        ws.cell(row=5, column=1, value="(none)")
    for i, (_, r) in enumerate(sub.iterrows(), start=5):
        status = r.get("Status Code")
        try:
            sc = int(float(status)) if pd.notna(status) else ""
        except (TypeError, ValueError):
            sc = ""
        # Operator-staged target values from page_execution take precedence
        # over the "[Phase 4 spec]" placeholder when present.
        def _override(col: str, fallback: str) -> str:
            v = r.get(col)
            if pd.notna(v) and v not in (None, ""):
                return v
            return fallback
        vals = [
            r["URL"], sc,
            int(num(r.get("Sessions"))),
            int(num(r.get("Referring Domains"))),
            r.get("Best TV Keyword") if pd.notna(r.get("Best TV Keyword")) else "",
            int(num(r.get("Best TV KW Rank"))) if pd.notna(r.get("Best TV KW Rank")) else "",
            int(num(r.get("Best TV KW SV"))) if pd.notna(r.get("Best TV KW SV")) else "",
            r["Logic"],
            "Restore the page. Users still arriving at this URL.",
            _override("Target H1", "[Phase 4 spec]"),
            _override("Target Title", "[Phase 4 spec]"),
            _override("Target Meta", "[Phase 4 spec]"),
        ]
        for j, v in enumerate(vals, start=1):
            ws.cell(row=i, column=j, value=v).alignment = Alignment(vertical="top", wrap_text=True)
    autosize(ws, 12, max_w=40)
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["H"].width = 60


# ─── service keyword heuristic ─────────────────────────────────────────────
def _services_from_domain(domain: str) -> tuple[str, ...]:
    """Pick a small set of URL-pattern keywords to feed classify_category.
    We're domain-agnostic in the platform, so use a generic transport-ish
    default set with light specialization based on the host."""
    d = (domain or "").lower()
    base = ["bus-hire", "charter", "minibus", "mini-bus", "coach"]
    if "party" in d:
        base += ["party-bus", "wedding", "hens", "bucks"]
    if "tour" in d:
        base += ["tour", "tours"]
    if "wedding" in d:
        base += ["wedding"]
    if "airport" in d or "transfer" in d:
        base += ["airport", "transfer"]
    return tuple(dict.fromkeys(base))


# ─── public entrypoint ─────────────────────────────────────────────────────
def build_phase1_workbook(
    bq_rows: list[dict],
    *,
    title: str,
    primary_host: str,
    domain: str,
    overrides: dict[str, dict] | None = None,
    executions: dict[str, dict] | None = None,
) -> BytesIO:
    """Compose the 12-tab Phase 1 WQA workbook in memory and return a
    BytesIO ready to stream as the HTTP response body.

    Parameters
    ----------
    bq_rows
        The 44-column rows returned by `/api/wqa/pages` (snake_case keys
        as they come out of `wqa_output`).
    title
        Site display name, used in tab titles (e.g. "BusCharter").
    primary_host
        Canonical origin used by redirect/canonicalization destination
        guesses, e.g. "https://www.buscharter.com.au".
    domain
        Bare primary domain (e.g. "buscharter.com.au"); seeds the
        services keyword list.
    overrides
        Optional map keyed by URL → {action, logic} from wqa_decision.
        When present, the override REPLACES the SOP-derived action and
        the logic string is prefixed with "[Override: <who>]" by the
        caller; this builder just consumes them as final values.
    executions
        Optional map keyed by URL → page_execution row dict. Used to
        seed Target H1 / Target Title / Target Meta / Target URL into
        the Restore URLs + Redirect Map tabs.
    """
    overrides = overrides or {}
    executions = executions or {}
    services = _services_from_domain(domain)

    # Build the title-cased DataFrame the rest of the pipeline expects.
    raw = pd.DataFrame(bq_rows)
    # Rename only columns that exist (BQ schema is stable but defensive).
    rename = {k: v for k, v in COLUMN_RENAME.items() if k in raw.columns}
    df = raw.rename(columns=rename).copy()
    # Ensure expected columns exist even if BQ has fewer.
    for col in COLUMN_RENAME.values():
        if col not in df.columns:
            df[col] = None

    df["Category"] = df["URL"].apply(lambda u: classify_category(u or "", services))
    loc = df["URL"].apply(
        lambda u: pd.Series(classify_location(u or ""), index=["Location Type", "Location"])
    )
    df["Location Type"] = loc["Location Type"]
    df["Location"] = loc["Location"]
    df["Service"] = df.apply(
        lambda r: classify_service(r["URL"] or "", r["Category"], services), axis=1
    )
    df["Primary?"] = df.apply(
        lambda r: is_primary(r.get("Status Code"), r.get("Indexability")), axis=1
    )

    # SOP-derived action first, then overlay overrides.
    actions = df.apply(assign_action, axis=1, result_type="expand")
    actions.columns = ["Action", "Logic"]
    df["Action"] = actions["Action"]
    df["Logic"] = actions["Logic"]

    if overrides:
        def _apply_override(row):
            o = overrides.get(row["URL"])
            if not o:
                return row
            row["Action"] = o.get("action", row["Action"])
            note = o.get("logic") or o.get("decided_by") or "human override"
            row["Logic"] = f"[Override: {note}] {row['Logic']}"
            return row
        df = df.apply(_apply_override, axis=1)

    # Stage target columns from page_execution onto each row so the
    # Restore + Redirect tabs can prefer operator values.
    if executions:
        target_url, target_h1, target_title, target_meta = [], [], [], []
        for url in df["URL"].tolist():
            ex = executions.get(url) or {}
            target_url.append(ex.get("target_url"))
            target_h1.append(ex.get("target_h1"))
            target_title.append(ex.get("target_title"))
            target_meta.append(ex.get("target_meta"))
        df["Target URL"] = target_url
        df["Target H1"] = target_h1
        df["Target Title"] = target_title
        df["Target Meta"] = target_meta

    counts = {
        "optimize": int(df["Action"].str.startswith("Optimize").sum()),
        "restore": int(df["Action"].str.startswith("Restore").sum()),
        "redirect": int(df["Action"].str.startswith("Redirect").sum()),
        "remove": int(df["Action"].str.startswith("Remove").sum()),
        "consolidate": int(df["Action"].str.startswith("Consolidate").sum()),
        "review": int(df["Action"].str.startswith("Review").sum()),
        "evaluate": int(df["Action"].str.startswith("Evaluate").sum()),
        "investigate": int(df["Action"].str.startswith("Investigate").sum()),
        "no_action": int(df["Action"].str.startswith("No Action").sum()),
    }

    action_plan_rows = [
        (1, f"Execute {counts['redirect']} redirects", counts["redirect"],
         "High" if counts["redirect"] else "Low", "Client Dev", "Redirect Map tab"),
        (2, f"Restore {counts['restore']} broken pages with material value", counts["restore"],
         "High" if counts["restore"] else "Low", "Client Dev + Skyward", "Restore URLs tab"),
        (3, f"Apply noindex / delete on {counts['remove']} zero-value URLs", counts["remove"],
         "Medium", "Skyward", "Removal List tab"),
        (4, f"Consolidate {counts['consolidate']} duplicate template pages", counts["consolidate"],
         "Medium" if counts["consolidate"] else "Low", "Skyward + Client Dev", "Canonicalization Map tab"),
        (5, f"Resolve {counts['review']} Review + {counts['evaluate']} Evaluate items",
         counts["review"] + counts["evaluate"], "Medium", "Skyward + Client", "URL Triage tab"),
        (6, "Audit canonicals on Optimize URLs", counts["optimize"], "Medium",
         "Skyward + Client Dev", "Canonical Audit tab"),
        (7, "Re-crawl with Screaming Frog after fixes", 0, "Low", "Skyward",
         "Implementation Checklist tab"),
        (8, f"Hand {counts['optimize'] + counts['restore']} URLs to Phase 2 + Phase 3",
         counts["optimize"] + counts["restore"], "High", "Skyward", "URL Optimization tab"),
    ]

    wb = Workbook()
    wb.remove(wb.active)
    write_action_legend(wb)
    write_action_plan(wb, action_plan_rows)
    write_url_triage(wb, df)
    write_funnel_summary(wb, df, total=len(df), title=title)
    write_service_summary(wb, df)
    write_redirect_map(wb, df, primary_host, title)
    write_canonicalization_map(wb, df, primary_host, title)
    write_canonical_audit(wb, df, title)
    write_removal_list(wb, df, title)
    write_implementation_checklist(wb, counts, title)
    write_url_optimization(wb, df, title)
    write_restore_urls(wb, df, title)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_phase1_dataframe(
    bq_rows: list[dict],
    *,
    domain: str,
    overrides: dict[str, dict] | None = None,
) -> pd.DataFrame:
    """Public helper that returns the triage DataFrame the Phase 2 builder
    consumes (instead of reading CSV). Keeps the column shape identical
    to the CSV that build_phase1_wqa.py emits."""
    overrides = overrides or {}
    services = _services_from_domain(domain)
    raw = pd.DataFrame(bq_rows)
    rename = {k: v for k, v in COLUMN_RENAME.items() if k in raw.columns}
    df = raw.rename(columns=rename).copy()
    for col in COLUMN_RENAME.values():
        if col not in df.columns:
            df[col] = None
    df["Category"] = df["URL"].apply(lambda u: classify_category(u or "", services))
    loc = df["URL"].apply(
        lambda u: pd.Series(classify_location(u or ""), index=["Location Type", "Location"])
    )
    df["Location Type"] = loc["Location Type"]
    df["Location"] = loc["Location"]
    df["Service"] = df.apply(
        lambda r: classify_service(r["URL"] or "", r["Category"], services), axis=1
    )
    df["Primary?"] = df.apply(
        lambda r: is_primary(r.get("Status Code"), r.get("Indexability")), axis=1
    )
    actions = df.apply(assign_action, axis=1, result_type="expand")
    actions.columns = ["Action", "Logic"]
    df["Action"] = actions["Action"]
    df["Logic"] = actions["Logic"]

    if overrides:
        def _apply_override(row):
            o = overrides.get(row["URL"])
            if not o:
                return row
            row["Action"] = o.get("action", row["Action"])
            note = o.get("logic") or o.get("decided_by") or "human override"
            row["Logic"] = f"[Override: {note}] {row['Logic']}"
            return row
        df = df.apply(_apply_override, axis=1)
    return df


__all__ = [
    "build_phase1_workbook",
    "build_phase1_dataframe",
    "COLUMN_RENAME",
]
