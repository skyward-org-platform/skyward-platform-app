"""Phase 4 Content Workbook builder — mirrors
`/Users/paulskirbe/agency/delivery/tna/build_phase4_content.py` but reads
from Supabase `content_row` + `keyword_cluster` instead of Phase 1 CSVs.

Public entrypoint: :func:`build_phase4_workbook`. Accepts the property
slug + a list of content_row dicts + a list of keyword_cluster dicts
(both as returned by the Supabase REST API). Returns a BytesIO with
the 2-tab workbook (Master Content Plan + Performance Tracker).

Visual output is byte-identical to the CLI builder so strategists can
swap an exported file against an existing CSV diff.
"""
from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ─── Styling (matches build_phase4_content.py) ─────────────────────────────
NAVY = "0F172A"
ORANGE = "F59E0B"
SLATE = "475569"
WHITE = "FFFFFF"
ALT = "F1F5F9"
THIN = Side(style="thin", color="E2E8F0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header(ws, row, ncols, bg=NAVY):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(name="Inter", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 36


def _style_body(ws, start_row, end_row, ncols):
    for r in range(start_row, end_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name="Inter", size=10, color="1E293B")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
            if (r - start_row) % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=ALT)


def _write_title(ws, title, subtitle, ncols):
    ws.cell(row=1, column=1, value=title).font = Font(name="Inter", size=16, bold=True, color=NAVY)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(row=2, column=1, value=subtitle).font = Font(name="Inter", size=10, italic=True, color=SLATE)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)


# ─── Header schema (must match build_phase4_content.py exactly) ────────────
HEADERS = [
    # Identity & Strategy (10)
    "URL", "Vertical", "Action Type", "Page Type", "Parent Page",
    "Priority Tier", "Primary Cluster", "Cluster SV", "Cluster KD", "Target Keyword",
    # Calendar (5)
    "Brief Due", "Draft Due", "Target Publish", "Owners", "Calendar Status",
    # Brief Spec (6)
    "Title (formatted)", "H1", "Meta Description Spec", "Word Count Target",
    "Phase 2 Yellow Resolution", "Brief Status",
    # Content Inputs (3)
    "Entities to Include", "FAQs to Answer", "Fan-Out Queries",
    # Draft & Production (6)
    "Writer", "Word Count Actual", "Draft Status", "Draft Link", "Published URL", "Feedback Notes",
    # Dependencies (1)
    "Dependencies",
    # Internal Linking (2)
    "Internal Links Out", "Internal Links In",
    # Schema (3)
    "Current Schema", "Required Schema", "JSON-LD Spec Notes",
    # Post-Publish (1)
    "Post-Publish Tasks (T+7 / T+30 / T+90)",
]
COL_WIDTHS = [
    50, 16, 12, 14, 22, 22, 32, 12, 10, 30,
    12, 12, 14, 32, 14,
    36, 26, 50, 16, 50, 13,
    50, 50, 45,
    16, 14, 14, 22, 30, 28,
    32,
    50, 50,
    24, 42, 50,
    55,
]
assert len(HEADERS) == len(COL_WIDTHS) == 37


# ─── Performance Tracker forecast helpers ──────────────────────────────────
def _target_rank_6mo(action: str) -> int:
    return 8 if action in ("Optimize", "Refresh") else 12


def _target_rank_12mo(action: str) -> int:
    return 4 if action in ("Optimize", "Refresh") else 6


def _est_6mo_clicks(sv: int, target_6mo: int) -> int:
    if not sv:
        return 0
    ctr = 0.07 if target_6mo <= 10 else 0.02
    return round(sv * ctr)


def _est_12mo_clicks(sv: int, target_12mo: int) -> int:
    if not sv:
        return 0
    ctr = 0.15 if target_12mo <= 5 else 0.07
    return round(sv * ctr)


# ─── Field fallback helpers ────────────────────────────────────────────────
def _coalesce(*vals):
    for v in vals:
        if v not in (None, ""):
            return v
    return None


def _action_type_effective(row: dict) -> str:
    return row.get("action_type_override") or row.get("action_type") or "Optimize"


def _title_effective(row: dict) -> str:
    return _coalesce(row.get("title_override"), row.get("title_formatted")) or ""


def _h1_effective(row: dict) -> str:
    return _coalesce(row.get("h1_override"), row.get("h1_target")) or ""


def _meta_effective(row: dict) -> str:
    return _coalesce(row.get("meta_description_override"), row.get("meta_description_spec")) or ""


# ─── Sort: priority tier (numeric prefix asc), then cluster SV desc ────────
def _priority_tier_num(tier: str | None) -> int:
    if not tier:
        return 99
    head = tier.split(".")[0].strip()
    try:
        return int(head)
    except ValueError:
        return 99


def _sort_rows(rows: list[dict], cluster_by_id: dict[str, dict]) -> list[dict]:
    def key(r):
        cluster = cluster_by_id.get(r.get("cluster_id") or "") or {}
        sv = cluster.get("total_sv") or 0
        return (_priority_tier_num(r.get("priority_tier")), -int(sv), r.get("url") or "")
    return sorted(rows, key=key)


# ─── Public entrypoint ─────────────────────────────────────────────────────
def build_phase4_workbook(
    *,
    rows: list[dict],
    clusters: list[dict],
    title: str,
    domain: str,
    sprint_start: str | None = None,
) -> BytesIO:
    """Build the 2-tab Phase 4 Content Workbook.

    Args:
        rows: list of content_row records from Supabase.
        clusters: list of keyword_cluster records (id, total_sv, avg_kd, head_term, name_override).
        title: property display name (used in tab title).
        domain: primary_domain (used in subtitle).
        sprint_start: ISO date of sprint 1 start, optional (subtitle only).

    Returns:
        BytesIO containing the xlsx bytes.
    """
    cluster_by_id: dict[str, dict] = {c["id"]: c for c in clusters if c.get("id")}
    sorted_rows = _sort_rows(rows, cluster_by_id)
    sprint_count = max(
        (r.get("sprint") or 0 for r in sorted_rows), default=0
    )

    wb = Workbook()
    wb.remove(wb.active)

    # ─── Tab 1: Master Content Plan ────────────────────────────────────────
    ws = wb.create_sheet("Master Content Plan")
    ws.sheet_properties.tabColor = NAVY

    subtitle_parts = [
        f"{domain}",
        f"{len(sorted_rows)} URLs",
    ]
    if sprint_count:
        subtitle_parts.append(f"{sprint_count} sprints")
    if sprint_start:
        subtitle_parts.append(f"starting {sprint_start}")
    subtitle = (
        " · ".join(subtitle_parts)
        + ". One row per URL. Content Inputs (entities/FAQs/fan-out) blocked this round, backfill per brief."
    )
    _write_title(ws, f"{title} — Master Content Plan", subtitle, len(HEADERS))

    ws.append([None])
    ws.append(HEADERS)
    hdr_row = ws.max_row
    _style_header(ws, hdr_row, len(HEADERS))

    for r in sorted_rows:
        cluster = cluster_by_id.get(r.get("cluster_id") or "") or {}
        primary_cluster = (
            cluster.get("name_override")
            or cluster.get("head_term")
            or "—"
        )
        cluster_sv = cluster.get("total_sv") or 0
        cluster_kd_raw = cluster.get("avg_kd")
        try:
            cluster_kd = float(cluster_kd_raw) if cluster_kd_raw is not None else None
        except (TypeError, ValueError):
            cluster_kd = None

        action_type = _action_type_effective(r)
        title_cell = _title_effective(r) or "(see brief notes)"
        h1_cell = _h1_effective(r) or "(see brief notes)"
        meta_cell = _meta_effective(r) or ""

        ws.append([
            r.get("url"),
            r.get("vertical") or "—",
            action_type,
            r.get("page_type") or "—",
            r.get("parent_page") or "—",
            r.get("priority_tier") or "—",
            primary_cluster,
            cluster_sv,
            cluster_kd,
            r.get("target_keyword") or "",
            r.get("brief_due"),
            r.get("draft_due"),
            r.get("target_publish"),
            r.get("owners") or "Skyward (writer) + Client (review)",
            r.get("calendar_status") or "Scheduled",
            title_cell,
            h1_cell,
            meta_cell,
            r.get("word_count_target") or "",
            r.get("phase2_yellow_resolution") or "",
            r.get("brief_status") or "Not Started",
            r.get("entities_blocked") or "",
            r.get("faqs_blocked") or "",
            r.get("fanout_blocked") or "",
            r.get("writer") or "TBD",
            r.get("word_count_actual"),
            r.get("status") or "Not Started",
            r.get("draft_link") or "",
            r.get("published_url") or "",
            r.get("feedback_notes") or "",
            r.get("dependencies") or "",
            r.get("internal_links_out") or "(none)",
            r.get("internal_links_in") or "(none)",
            r.get("current_schema") or "—",
            r.get("required_schema") or "",
            r.get("jsonld_notes") or "",
            r.get("post_publish_tasks") or "",
        ])

    end_row = ws.max_row
    if end_row > hdr_row:
        _style_body(ws, hdr_row + 1, end_row, len(HEADERS))
    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for r in range(hdr_row + 1, end_row + 1):
        ws.row_dimensions[r].height = 60
    ws.freeze_panes = ws.cell(row=hdr_row + 1, column=2)
    if end_row > hdr_row:
        ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(len(HEADERS))}{end_row}"

    # ─── Tab 2: Performance Tracker ────────────────────────────────────────
    ws2 = wb.create_sheet("Performance Tracker")
    ws2.sheet_properties.tabColor = ORANGE
    pt_headers = [
        "URL", "Target Keyword", "Cluster SV", "Action Type",
        "6mo Target Rank", "6mo Est. Clicks", "12mo Target Rank", "12mo Est. Clicks",
        "30d Rank", "60d Rank", "90d Rank", "Refresh Flag?",
    ]
    tracker_rows = [r for r in sorted_rows if (r.get("target_keyword") or "").strip()]
    _write_title(
        ws2,
        f"{title} — Performance Tracker",
        (
            f"{len(tracker_rows)} URLs with a target keyword. Tracked at 30/60/90 days vs Phase 3 forecast. "
            "Refresh flag fires when 90-day rank is 30%+ worse than the 6-month target. "
            "6mo + 12mo projections use SOP CTR curves (top 10 = 7%, top 5 = 15%)."
        ),
        len(pt_headers),
    )
    ws2.append([None])
    ws2.append(pt_headers)
    hdr2 = ws2.max_row
    _style_header(ws2, hdr2, len(pt_headers))

    for r in tracker_rows:
        action = _action_type_effective(r)
        cluster = cluster_by_id.get(r.get("cluster_id") or "") or {}
        sv = int(cluster.get("total_sv") or 0)
        t6 = _target_rank_6mo(action)
        t12 = _target_rank_12mo(action)
        e6 = _est_6mo_clicks(sv, t6)
        e12 = _est_12mo_clicks(sv, t12)

        r90 = r.get("rank_90d")
        refresh_flag = "Yes" if (r90 is not None and r90 > t6 * 1.43) else None

        ws2.append([
            r.get("url"),
            r.get("target_keyword"),
            sv,
            action,
            t6, e6, t12, e12,
            r.get("rank_30d"),
            r.get("rank_60d"),
            r.get("rank_90d"),
            refresh_flag,
        ])

    end_row2 = ws2.max_row
    if end_row2 > hdr2:
        _style_body(ws2, hdr2 + 1, end_row2, len(pt_headers))
    pt_widths = [50, 30, 12, 14, 14, 14, 14, 14, 10, 10, 10, 14]
    for i, w in enumerate(pt_widths, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    for r in range(hdr2 + 1, end_row2 + 1):
        ws2.row_dimensions[r].height = 22
    ws2.freeze_panes = ws2.cell(row=hdr2 + 1, column=2)
    if end_row2 > hdr2:
        ws2.auto_filter.ref = f"A{hdr2}:{get_column_letter(len(pt_headers))}{end_row2}"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
